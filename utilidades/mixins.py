"""
Mixins reutilizables para ViewSets de DRF.
"""
from io import BytesIO

from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, inline_serializer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import serializers
from rest_framework.decorators import action

from utilidades.filtros import aplicar_filtros, aplicar_ordenamientos

class _FiltroItemSerializer(serializers.Serializer):
    propiedad = serializers.CharField()
    operador = serializers.CharField(
        help_text='=, !=, >, >=, <, <=, contiene, comienza_con, termina_con, in, is_null',
    )
    valor = serializers.JSONField(required=False, allow_null=True)
    operador_logico = serializers.ChoiceField(choices=['AND', 'OR'], required=False)


_BusquedaRequest = inline_serializer(
    name='BusquedaRequest',
    fields={
        'filtros': _FiltroItemSerializer(many=True, required=False),
        'ordenamientos': serializers.ListField(
            child=serializers.CharField(), required=False,
            help_text='Lista de campos. Prefijo "-" para descendente.',
        ),
    },
)


class FiltrosDinamicosMixin:
    """
    Agrega la acción `POST /<recurso>/lista/` con filtros dinámicos y paginación.

    El ViewSet que lo herede debe declarar:
        campos_filtrables: set[str]                   # whitelist obligatoria

    Y opcionalmente:
        select_related_lista: tuple[str, ...]         # FKs a precargar
        prefetch_related_lista: tuple[str, ...]       # M2M / reverse a precargar
        ordenamiento_default_lista: tuple[str, ...]   # si no envían `ordenamientos`

    Para casos especiales (filtrado forzado por tenant/usuario, annotations, etc.)
    sobreescribe `get_queryset_lista`.
    """

    campos_filtrables: set = set()
    select_related_lista: tuple = ()
    prefetch_related_lista: tuple = ()
    ordenamiento_default_lista: tuple = ()

    def get_queryset_lista(self):
        qs = self.get_queryset()
        if self.select_related_lista:
            qs = qs.select_related(*self.select_related_lista)
        if self.prefetch_related_lista:
            qs = qs.prefetch_related(*self.prefetch_related_lista)
        return qs

    @extend_schema(
        summary='Listar con filtros dinámicos',
        description=(
            'Lista registros aplicando filtros dinámicos enviados en el body. '
            'Operadores soportados: =, !=, >, >=, <, <=, contiene, comienza_con, '
            'termina_con, in, is_null. Combinación AND/OR vía `operador_logico` '
            'por filtro (AND por defecto, evaluación secuencial).'
        ),
        request=_BusquedaRequest,
    )
    @action(detail=False, methods=['post'])
    def lista(self, request):
        filtros = request.data.get('filtros') or []
        ordenamientos = request.data.get('ordenamientos') or []

        qs = self.get_queryset_lista()
        qs = aplicar_filtros(qs, filtros, self.campos_filtrables)
        qs = aplicar_ordenamientos(qs, ordenamientos, self.campos_filtrables)
        if not ordenamientos and self.ordenamiento_default_lista:
            qs = qs.order_by(*self.ordenamiento_default_lista)

        pagina = self.paginate_queryset(qs)
        serializador = self.get_serializer(pagina, many=True)
        return self.get_paginated_response(serializador.data)


class ExcelMixin:
    """
    Agrega la acción `POST /<recurso>/excel/` que devuelve un archivo .xlsx
    con los datos filtrados (mismos filtros/ordenamientos que `lista`, sin paginar).

    El ViewSet que lo herede debe declarar:
        campos_excel: tuple[tuple[str, str], ...]
            Pares (campo, encabezado). `campo` admite notación dotted para
            relaciones (p.ej. 'ciudad.nombre'). El encabezado es el título de la columna.

    Y opcionalmente:
        nombre_archivo_excel: str   # nombre base sin extensión (default: nombre del modelo)
        hoja_excel: str             # nombre de la hoja (default: 'Datos')

    Requiere también que el ViewSet defina `campos_filtrables` (igual que
    `FiltrosDinamicosMixin`) si se quieren aplicar filtros.
    """

    campos_excel: tuple = ()
    nombre_archivo_excel: str = ''
    hoja_excel: str = 'Datos'

    def get_queryset_excel(self):
        if hasattr(self, 'get_queryset_lista'):
            return self.get_queryset_lista()
        return self.get_queryset()

    @extend_schema(
        summary='Exportar a Excel',
        description=(
            'Devuelve un archivo .xlsx con los registros filtrados. '
            'Acepta los mismos `filtros` y `ordenamientos` que la acción `lista`.'
        ),
        request=_BusquedaRequest,
        responses={
            (200, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'): bytes,
        },
    )
    @action(detail=False, methods=['post'])
    def excel(self, request):
        filtros = request.data.get('filtros') or []
        ordenamientos = request.data.get('ordenamientos') or []

        qs = self.get_queryset_excel()
        campos_filtrables = getattr(self, 'campos_filtrables', set())
        qs = aplicar_filtros(qs, filtros, campos_filtrables)
        qs = aplicar_ordenamientos(qs, ordenamientos, campos_filtrables)
        if not ordenamientos:
            default = getattr(self, 'ordenamiento_default_lista', ())
            if default:
                qs = qs.order_by(*default)

        wb = Workbook()
        ws = wb.active
        ws.title = self.hoja_excel

        fuente_encabezado = Font(bold=True)
        fondo_encabezado = PatternFill('solid', fgColor='D9D9D9')
        for col, (_, encabezado) in enumerate(self.campos_excel, start=1):
            celda = ws.cell(row=1, column=col, value=encabezado)
            celda.font = fuente_encabezado
            celda.fill = fondo_encabezado

        for fila, obj in enumerate(qs.iterator(), start=2):
            for col, (campo, _) in enumerate(self.campos_excel, start=1):
                ws.cell(row=fila, column=col, value=self._valor_excel(obj, campo))

        for col, (campo, encabezado) in enumerate(self.campos_excel, start=1):
            ancho = max(len(str(encabezado)), 12)
            ws.column_dimensions[get_column_letter(col)].width = min(ancho + 2, 50)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        nombre = self.nombre_archivo_excel or qs.model._meta.model_name
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre}.xlsx"'
        return response

    @staticmethod
    def _valor_excel(obj, campo):
        valor = obj
        for parte in campo.split('.'):
            if valor is None:
                return None
            valor = getattr(valor, parte, None)
        if isinstance(valor, bool):
            return 'Sí' if valor else 'No'
        return valor
