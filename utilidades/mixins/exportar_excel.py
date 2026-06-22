"""
Mixin para exportar datos de un ViewSet a un archivo .xlsx.
"""
from io import BytesIO

from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.decorators import action

from utilidades.filtros import aplicar_filtros, aplicar_ordenamientos
from utilidades.mixins.filtros import BusquedaRequest

# Fuente cross-platform (Arial está en Windows/Mac y en Linux con msttcorefonts).
# El default de openpyxl es Calibri, que no viene en LibreOffice por defecto.
_FUENTE = 'Arial'
_FUENTE_TAM = 10
_FUENTE_NORMAL = Font(name=_FUENTE, size=_FUENTE_TAM)
_FUENTE_ENCABEZADO = Font(name=_FUENTE, size=_FUENTE_TAM, bold=True)


def _crear_workbook():
    wb = Workbook()
    wb._fonts[0] = _FUENTE_NORMAL
    return wb


class ExportarExcelMixin:
    """
    Agrega la acción `POST /<recurso>/excel/` que devuelve un archivo .xlsx
    con los datos filtrados (mismos filtros/ordenamientos que `lista`, sin paginar).

    El ViewSet que lo herede debe declarar:
        serializer_class_exportar: Serializer
            Serializer que define la estructura del Excel. Debe exponer:
                model:          clase del modelo Django
                campos_excel:   tuple[tuple[campo, encabezado], ...]
                valor_excel(obj, campo) -> Any
                nombre_archivo: str (opcional; default = nombre del modelo)
                hoja:           str (opcional; default = 'Datos')
    """

    serializer_class_exportar = None

    def get_serializer_exportar(self):
        if self.serializer_class_exportar is None:
            raise NotImplementedError(
                f"{type(self).__name__} debe declarar `serializer_class_exportar`"
            )
        return self.serializer_class_exportar()

    def get_queryset_excel(self):
        if hasattr(self, 'get_queryset_lista'):
            return self.get_queryset_lista()
        return self.get_queryset()

    def _config_lista(self, name, default):
        """Lee config (campos_filtrables, ordenamiento_default_lista) del serializer principal o del viewset."""
        serializer_cls = self.get_serializer_class()
        if hasattr(serializer_cls, name):
            return getattr(serializer_cls, name)
        return getattr(self, name, default)

    @extend_schema(
        summary='Exportar a Excel',
        description=(
            'Devuelve un archivo .xlsx con los registros filtrados. '
            'Acepta los mismos `filtros` y `ordenamientos` que la acción `lista`.'
        ),
        request=BusquedaRequest,
        responses={
            (200, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'): bytes,
        },
    )
    @action(detail=False, methods=['post'])
    def excel(self, request):
        serializer = self.get_serializer_exportar()
        filtros = request.data.get('filtros') or []
        ordenamientos = request.data.get('ordenamientos') or []

        qs = self.get_queryset_excel()
        campos_filtrables = self._config_lista('campos_filtrables', set())
        qs = aplicar_filtros(qs, filtros, campos_filtrables)
        qs = aplicar_ordenamientos(qs, ordenamientos, campos_filtrables)
        if not ordenamientos:
            default = self._config_lista('ordenamiento_default_lista', ())
            if default:
                qs = qs.order_by(*default)

        wb = _crear_workbook()
        ws = wb.active
        ws.title = getattr(serializer, 'hoja', None) or 'Datos'

        campos = serializer.campos_excel
        fondo_encabezado = PatternFill('solid', fgColor='D9D9D9')
        for col, (_, encabezado) in enumerate(campos, start=1):
            celda = ws.cell(row=1, column=col, value=encabezado)
            celda.font = _FUENTE_ENCABEZADO
            celda.fill = fondo_encabezado

        # chunk_size es obligatorio en .iterator() cuando el queryset trae
        # prefetch_related (Django lo exige para acotar la memoria de los prefetch).
        for fila, obj in enumerate(qs.iterator(chunk_size=2000), start=2):
            for col, (campo, _) in enumerate(campos, start=1):
                celda = ws.cell(row=fila, column=col, value=serializer.valor_excel(obj, campo))
                celda.font = _FUENTE_NORMAL

        for col, (_, encabezado) in enumerate(campos, start=1):
            ancho = max(len(str(encabezado)), 12)
            ws.column_dimensions[get_column_letter(col)].width = min(ancho + 2, 50)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        nombre = getattr(serializer, 'nombre_archivo', None) or qs.model._meta.model_name
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre}.xlsx"'
        return response
