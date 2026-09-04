"""
Mixin para importar datos a un ViewSet desde un archivo .xlsx.
"""
from io import BytesIO
from zipfile import BadZipFile

from django.db import models as dj_models, transaction
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, inline_serializer
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from rest_framework import serializers, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response

from utilidades.throttles import ImportarUsuarioTenantThrottle

# Fuente cross-platform (Arial está en Windows/Mac y en Linux con msttcorefonts).
# El default de openpyxl es Calibri, que no viene en LibreOffice por defecto.
_FUENTE = 'Arial'
_FUENTE_TAM = 10
_FUENTE_NORMAL = Font(name=_FUENTE, size=_FUENTE_TAM)
_FUENTE_ENCABEZADO = Font(name=_FUENTE, size=_FUENTE_TAM, bold=True)


def _crear_workbook():
    wb = Workbook()
    wb._fonts[0] = _FUENTE_NORMAL
    return wb


_ImportarRequest = inline_serializer(
    name='ImportarRequest',
    fields={'archivo': serializers.FileField(help_text='Archivo .xlsx')},
)


class ImportarExcelMixin:
    """
    Agrega dos acciones al ViewSet:
        GET  /<recurso>/importar-ejemplo/  — descarga plantilla .xlsx
        POST /<recurso>/importar/          — sube archivo (multipart, campo `archivo`)

    El ViewSet que lo herede debe declarar:
        serializer_class_importar: Serializer
            Serializer que define la estructura del Excel y la lógica de creación.
            Debe exponer:
                model:          clase del modelo Django
                campos_excel:   tuple[tuple[campo, encabezado], ...]
                procesar_lote(filas_validas) -> (creados, errores)
                                # filas_validas: list[(fila, datos)] ya validadas
                                # estructuralmente. Devuelve cuántos creó y la
                                # lista [{fila, mensaje}] de lo que falló; si
                                # devuelve errores, el mixin revierte todo.
                                # El import solo crea, no actualiza: un registro
                                # que ya existe debe reportarse como error.
            Opcionales:
                campos_requeridos: set[str]
                nombre_archivo:    str
                valores_ejemplo:   dict[campo, (fila1, fila2)]
                errores_completos: bool  # reportar las dos fases de una vez,
                                         # y un error por problema en vez de
                                         # uno por fila

        Si el Excel depende de algo más que la clase (p.ej. el documento padre),
        el ViewSet puede sobrescribir `get_serializer_importar()` y construir el
        serializer a mano; el resto del flujo no cambia.

    Si CUALQUIER fila falla, toda la transacción se revierte y se devuelve 400
    con la lista de errores. La respuesta exitosa es `{creados: N}`.

    Convención FK: campos con notación dotted (p.ej. `ciudad.id`) se marcan en la
    plantilla con fondo amarillo y con el sufijo de la llave que se espera —`(ID)`
    para `ciudad.id`, `(Código)` para `cuenta.codigo`— para que el usuario sepa qué
    escribir. Ver `ETIQUETAS_LLAVE_FK`.
    """

    EXTENSIONES_VALIDAS_IMPORTAR = ('.xlsx',)
    LIMITE_ERRORES_IMPORTAR = 100
    MAX_FILAS_IMPORTAR = 10_000
    MAX_TAMANO_ARCHIVO_BYTES = 5 * 1024 * 1024  # 5 MB
    serializer_class_importar = None

    def get_serializer_importar(self):
        if self.serializer_class_importar is None:
            raise NotImplementedError(
                f"{type(self).__name__} debe declarar `serializer_class_importar`"
            )
        return self.serializer_class_importar()

    # Cómo se rotula la llave de una columna FK. Casi todas las columnas dotted del
    # proyecto apuntan al PK (`cuenta.id`), pero un importador puede resolver por una
    # llave natural (`cuenta.codigo`) y el encabezado tiene que decir cuál se espera:
    # un usuario que ve "(ID)" escribe el id.
    ETIQUETAS_LLAVE_FK = {
        'id': 'ID',
        'codigo': 'Código',
        'numero_identificacion': 'Número identificación',
    }

    @classmethod
    def _encabezado_importar(cls, campo, encabezado, requeridos=()):
        sufijo = ''
        if '.' in campo:
            llave = campo.rsplit('.', 1)[1]
            sufijo += f' ({cls.ETIQUETAS_LLAVE_FK.get(llave, llave)})'
        if campo in requeridos:
            sufijo += ' *'
        return f'{encabezado}{sufijo}'

    @staticmethod
    def _error_importar(detail, *, status_code=status.HTTP_400_BAD_REQUEST, fase=None, errores=None):
        """
        Construye la respuesta de error unificada de importación.

        Shape estándar (todas las respuestas de error lo cumplen):
            {
                "detail": str,                      # siempre
                "fase": str,                        # opcional: encabezados | estructural |
                                                    #           negocio | validacion
                "total_errores": int,               # presente solo si hay `errores`
                "errores": [{"fila"?: int, "mensaje": str, "fase"?: str}]  # opcional
            }

        `validacion` es la fase mixta: aparece solo cuando el serializer declara
        `errores_completos` y el archivo falló en las dos fases a la vez. Ahí cada
        error trae su propia `fase`.
        """
        cuerpo = {'detail': detail}
        if fase is not None:
            cuerpo['fase'] = fase
        if errores is not None:
            cuerpo['total_errores'] = len(errores)
            cuerpo['errores'] = errores
        return Response(cuerpo, status=status_code)

    def _agregar_errores(self, destino, fila, mensajes, separados):
        """
        Agrega los problemas de una fila y dice si ya se llegó al tope.

        Con `separados` va uno por problema aunque la fila se repita, que es lo
        que deja ver todo lo que hay que arreglar; sin él van juntos en un solo
        mensaje, que es como responden los importadores de tabla plana.
        """
        if separados:
            destino.extend({'fila': fila, 'mensaje': m} for m in mensajes)
        else:
            destino.append({'fila': fila, 'mensaje': '; '.join(mensajes)})
        return len(destino) >= self.LIMITE_ERRORES_IMPORTAR

    @extend_schema(
        summary='Descargar plantilla de importación',
        description=(
            'Devuelve un archivo .xlsx con dos filas de ejemplo. Los campos FK van '
            'con fondo amarillo y el sufijo de la llave que esperan —"(ID)", '
            '"(Código)"—, que es lo que debe escribirse en la celda.'
        ),
    )
    @action(detail=False, methods=['get'], url_path='importar-ejemplo')
    def importar_ejemplo(self, request):
        serializer = self.get_serializer_importar()
        modelo = serializer.model

        wb = _crear_workbook()
        ws = wb.active
        ws.title = 'Datos'

        campos = serializer.campos_excel
        requeridos = getattr(serializer, 'campos_requeridos', set())
        fondo_normal = PatternFill('solid', fgColor='D9D9D9')
        fondo_fk = PatternFill('solid', fgColor='FFF2CC')  # amarillo claro = campo FK
        for col, (campo, encabezado) in enumerate(campos, start=1):
            es_fk = '.' in campo
            texto = self._encabezado_importar(campo, encabezado, requeridos)
            celda = ws.cell(row=1, column=col, value=texto)
            celda.font = _FUENTE_ENCABEZADO
            celda.fill = fondo_fk if es_fk else fondo_normal
            ancho = max(len(str(texto)), 12)
            ws.column_dimensions[get_column_letter(col)].width = min(ancho + 2, 50)

        # `valores_ejemplo` es opcional: solo hace falta cuando el valor derivado del
        # campo del modelo no orienta al usuario (columnas que no son un campo, como
        # una lista de ids, o campos con un dominio cerrado como D/C).
        valores_ejemplo = getattr(serializer, 'valores_ejemplo', None) or {}
        for fila in (1, 2):
            for col, (campo, _) in enumerate(campos, start=1):
                if campo in valores_ejemplo:
                    valor = valores_ejemplo[campo][fila - 1]
                else:
                    field = self._resolver_campo(modelo, campo)
                    valor = self._valor_ejemplo(field, fila, campo)
                celda = ws.cell(row=fila + 1, column=col, value=valor)
                celda.font = _FUENTE_NORMAL

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        nombre = getattr(serializer, 'nombre_archivo', None) or modelo._meta.model_name
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="importar_{nombre}_ejemplo.xlsx"'
        return response

    @staticmethod
    def _resolver_campo(modelo, dotted_path):
        partes = dotted_path.split('.')
        actual = modelo
        for parte in partes[:-1]:
            try:
                f = actual._meta.get_field(parte)
            except Exception:
                return None
            if not (f.is_relation and f.related_model):
                return None
            actual = f.related_model
        try:
            return actual._meta.get_field(partes[-1])
        except Exception:
            return None

    @staticmethod
    def _validar_tipo(field, valor, encabezado):
        """Devuelve mensaje de error si `valor` no coincide con el tipo del campo, sino None."""
        if field is None or valor in (None, ''):
            return None
        if isinstance(field, dj_models.BooleanField):
            if isinstance(valor, bool):
                return None
            if str(valor).strip().lower() in (
                'sí', 'si', 'no', 'true', 'false', '0', '1', 'yes', 'verdadero', 'falso',
            ):
                return None
            return f'{encabezado} debe ser Sí o No (recibido: "{valor}")'
        if isinstance(field, dj_models.DateField):
            import datetime as _dt
            if isinstance(valor, (_dt.date, _dt.datetime)):
                return None
            return f'{encabezado} debe ser una fecha (recibido: "{valor}")'
        if isinstance(field, dj_models.DecimalField):
            try:
                float(valor)
                return None
            except (TypeError, ValueError):
                return f'{encabezado} debe ser un número decimal (recibido: "{valor}")'
        if isinstance(field, (
            dj_models.BigIntegerField, dj_models.IntegerField,
            dj_models.SmallIntegerField, dj_models.PositiveIntegerField,
            dj_models.AutoField,
        )):
            try:
                int(valor)
                return None
            except (TypeError, ValueError):
                return f'{encabezado} debe ser un número entero (recibido: "{valor}")'
        return None  # CharField, TextField, etc. aceptan cualquier valor

    @staticmethod
    def _valor_ejemplo(field, fila, campo):
        if campo == 'id':
            return None
        if field is None:
            return f'ejemplo {fila}'
        if isinstance(field, dj_models.BooleanField):
            return 'Sí' if fila == 1 else 'No'
        if isinstance(field, dj_models.DateField):
            return '2026-01-15' if fila == 1 else '2026-12-31'
        if isinstance(field, dj_models.DecimalField):
            return 100 * fila
        if isinstance(field, (
            dj_models.BigIntegerField, dj_models.IntegerField,
            dj_models.SmallIntegerField, dj_models.PositiveIntegerField,
            dj_models.AutoField,
        )):
            return fila
        return f'ejemplo {fila}'

    @extend_schema(
        summary='Importar desde Excel',
        description=(
            'Recibe un archivo .xlsx en el campo `archivo` (multipart/form-data). '
            'La primera fila debe contener los encabezados de la plantilla. '
            'Procesamiento todo-o-nada: si alguna fila falla, no se guarda nada.'
        ),
        request={'multipart/form-data': _ImportarRequest},
    )
    @action(
        detail=False, methods=['post'],
        parser_classes=[MultiPartParser],
        throttle_classes=[ImportarUsuarioTenantThrottle],
    )
    def importar(self, request):
        serializer = self.get_serializer_importar()
        modelo = serializer.model

        archivo = request.FILES.get('archivo')
        if not archivo:
            return self._error_importar("Debe enviar el archivo en el campo 'archivo'")

        nombre = archivo.name.lower()
        if not any(nombre.endswith(ext) for ext in self.EXTENSIONES_VALIDAS_IMPORTAR):
            return self._error_importar(
                f"Extensión no permitida. Solo se aceptan: "
                f"{', '.join(self.EXTENSIONES_VALIDAS_IMPORTAR)}"
            )

        if archivo.size > self.MAX_TAMANO_ARCHIVO_BYTES:
            mb = self.MAX_TAMANO_ARCHIVO_BYTES // (1024 * 1024)
            return self._error_importar(f'El archivo supera el tamaño máximo permitido ({mb} MB)')

        try:
            # data_only=False permite detectar fórmulas vía cell.data_type == 'f'
            wb = load_workbook(archivo, data_only=False, read_only=True)
        except (InvalidFileException, BadZipFile):
            return self._error_importar('El archivo no es un Excel válido o está corrupto')
        except Exception as e:
            return self._error_importar(f'No se pudo leer el archivo: {e}')

        try:
            ws = wb.active
            rows = ws.iter_rows()  # cell objects (necesario para detectar fórmulas)

            try:
                header_row = next(rows)
            except StopIteration:
                return self._error_importar('El archivo no tiene contenido')
            headers_archivo = [c.value for c in header_row]

            campos = serializer.campos_excel
            requeridos = getattr(serializer, 'campos_requeridos', set())
            esperados = [self._encabezado_importar(c, e, requeridos) for c, e in campos]
            recibidos = [h for h in headers_archivo if h is not None]
            # El orden de las columnas no importa: cada celda se mapea por nombre de
            # encabezado (ver `mapping`). Solo se exige que estén exactamente las mismas.
            if set(recibidos) != set(esperados):
                errores_encabezados = (
                    [{'mensaje': f'Falta la columna: {h}'} for h in esperados if h not in recibidos]
                    + [{'mensaje': f'Columna no reconocida: {h}'} for h in recibidos if h not in esperados]
                )
                return self._error_importar(
                    'Los encabezados del archivo no coinciden con la plantilla',
                    fase='encabezados',
                    errores=errores_encabezados,
                )

            mapping = {self._encabezado_importar(c, e, requeridos): c for c, e in campos}

            # ============ FASE 1: validación estructural (sin BD) ============
            # fórmulas, tipos de datos, campos requeridos.
            errores_completos = getattr(serializer, 'errores_completos', False)
            filas_validas = []
            errores_estructurales = []
            filas_procesadas = 0
            exceso_filas = False

            for idx, row in enumerate(rows, start=2):
                datos = {}
                formulas = []
                for col, header in enumerate(headers_archivo):
                    campo = mapping.get(header)
                    if not campo:
                        continue
                    celda = row[col] if col < len(row) else None
                    if celda is not None and celda.data_type == 'f':
                        formulas.append(header)
                        datos[campo] = None
                    else:
                        datos[campo] = celda.value if celda is not None else None

                if not any(v not in (None, '') for v in datos.values()) and not formulas:
                    continue
                filas_procesadas += 1

                if filas_procesadas > self.MAX_FILAS_IMPORTAR:
                    exceso_filas = True
                    break

                if formulas:
                    errores_estructurales.append({
                        'fila': idx,
                        'mensaje': f'Contiene fórmulas no permitidas en: {", ".join(formulas)}',
                    })
                    if len(errores_estructurales) >= self.LIMITE_ERRORES_IMPORTAR:
                        break
                    continue

                errores_tipo = []
                for campo, encabezado in campos:
                    msg = self._validar_tipo(
                        self._resolver_campo(modelo, campo),
                        datos.get(campo),
                        encabezado,
                    )
                    if msg:
                        errores_tipo.append(msg)
                if errores_tipo:
                    if self._agregar_errores(
                        errores_estructurales, idx, errores_tipo, errores_completos,
                    ):
                        break
                    continue

                faltantes = [
                    f'Falta el campo requerido: {encabezado}'
                    for campo, encabezado in campos
                    if campo in requeridos and datos.get(campo) in (None, '')
                ]
                if faltantes:
                    if self._agregar_errores(
                        errores_estructurales, idx, faltantes, errores_completos,
                    ):
                        break
                    continue

                filas_validas.append((idx, datos))
        finally:
            wb.close()

        if exceso_filas:
            return self._error_importar(
                f'El archivo supera el máximo de {self.MAX_FILAS_IMPORTAR} filas permitidas'
            )

        if filas_procesadas == 0:
            return self._error_importar('El archivo no tiene filas para importar')

        # Por defecto la fase estructural corta: no vale la pena consultar la BD por
        # un archivo que ya se sabe malo. Un importador puede pedir el reporte
        # completo (`errores_completos`) y entonces la fase 2 corre igual, solo para
        # recolectar sus errores: el usuario corrige el archivo una vez y no dos.
        errores_completos = getattr(serializer, 'errores_completos', False)
        if errores_estructurales and not errores_completos:
            return self._respuesta_errores_importar('estructural', errores_estructurales)

        # ============ FASE 2: lógica de negocio + creación (transaccional) ============
        # El serializer procesa el lote completo: pre-carga FKs, valida, bulk_create.
        with transaction.atomic():
            creados, errores_negocio = serializer.procesar_lote(filas_validas)
            if errores_negocio or errores_estructurales:
                transaction.set_rollback(True)

        # `validacion` es la fase mixta: si solo falló una, se reporta como siempre.
        if errores_estructurales and errores_negocio:
            return self._respuesta_errores_importar(
                'validacion', self._unir_errores(errores_estructurales, errores_negocio),
            )
        if errores_estructurales:
            return self._respuesta_errores_importar('estructural', errores_estructurales)
        if errores_negocio:
            return self._respuesta_errores_importar('negocio', errores_negocio)

        return Response({'creados': creados})

    def _unir_errores(self, estructurales, negocio):
        """
        Junta las dos fases en una sola lista ordenada por fila, con la fase en cada
        error. Se recorta al mismo tope que cada fase por separado.
        """
        unidos = (
            [{**e, 'fase': 'estructural'} for e in estructurales]
            + [{**e, 'fase': 'negocio'} for e in negocio]
        )
        unidos.sort(key=lambda e: e.get('fila') or 0)
        return unidos[:self.LIMITE_ERRORES_IMPORTAR]

    def _respuesta_errores_importar(self, fase, errores):
        limite_alcanzado = len(errores) >= self.LIMITE_ERRORES_IMPORTAR
        if limite_alcanzado:
            detail = (
                f'Se detectaron al menos {self.LIMITE_ERRORES_IMPORTAR} errores en la fase '
                f'{fase}. Se detuvo la validación. No se importó nada — corrija y reintente.'
            )
        elif fase == 'estructural':
            detail = (
                'El archivo tiene problemas estructurales (fórmulas, tipos o requeridos). '
                'No se procesó ningún registro.'
            )
        elif fase == 'validacion':
            detail = (
                'El archivo tiene errores de estructura y de datos. Cada error trae su '
                'fase. No se importó nada — corrija y reintente.'
            )
        else:
            detail = (
                'Errores de datos al intentar guardar (FK inexistente, duplicados, etc.). '
                'No se importó nada.'
            )
        return self._error_importar(detail, fase=fase, errores=errores)
