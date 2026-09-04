import io
import json
import uuid as uuid_lib
import zipfile
from datetime import date, time
from decimal import Decimal
from unittest import mock

import httpx
from botocore.exceptions import ClientError, ConnectionClosedError
from django.apps import apps
from django.conf import settings
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, connection, transaction
from django.test import SimpleTestCase
from django.test.utils import CaptureQueriesContext
from django.utils import timezone
from django_tenants.test.cases import TenantTestCase
from rest_framework import permissions
from rest_framework.exceptions import ValidationError
from rest_framework.test import APIRequestFactory, force_authenticate

from contabilidad.models import ConCentroCosto, ConCuenta
from general.models import (
    GenArchivo,
    GenArchivoTipo,
    GenAsesor,
    GenConfiguracion,
    GenDocumento,
    GenDocumentoDetalle,
    GenDocumentoTipo,
    GenFestivo,
    GenContacto,
    GenImpuesto,
    GenCiudad,
    GenEstado,
    GenIdentificacion,
    GenItem,
    GenModelo,
    GenPais,
    GenParametro,
    GenPrecio,
    GenPrecioDetalle,
    GenTipoPersona,
)
from general.servicios import documento as documento_servicio
from general.servicios import factura_electronica
from general.servicios import rededoc as rededoc_servicio
from general.serializers import (
    GenAsesorImportarSerializer,
    GenDocumentoDetalleImportarSerializer,
    GenParametroSerializer,
    GenPrecioDetalleImportarSerializer,
)
from general.views.archivo import GenArchivoViewSet
from general.views.configuracion import GenConfiguracionViewSet
from general.views.documento_detalle import GenDocumentoDetalleViewSet
from general.views.factura_electronica import GenFacturaElectronicaViewSet
from general.views.parametro import GenParametroViewSet
from general.views.modelo import GenModeloViewSet
from general.views.precio_detalle import GenPrecioDetalleViewSet
from seguridad.models import SegUsuario
from general.servicios import archivo as archivo_servicio
from utilidades import backblaze, mime
from utilidades.mixins import ImportarExcelMixin


class _ModeloViewSinPermisos(GenModeloViewSet):
    """Variante sin auth ni throttle: el usuario se inyecta con force_authenticate."""
    authentication_classes = []
    permission_classes = [permissions.AllowAny]
    throttle_classes = []


class _ArchivoViewSinPermisos(GenArchivoViewSet):
    """Misma idea que arriba: acá se prueba la validación, no la membresía."""
    authentication_classes = []
    permission_classes = [permissions.AllowAny]
    throttle_classes = []


class GenerarDocumentoTests(TenantTestCase):
    """
    Generación mensual: los detalles del destino se acotan a la ventana del mes y
    los que no se solapan con él no se generan.
    """

    @classmethod
    def setup_tenant(cls, tenant):
        tenant.nombre = 'Test'
        tenant.celular = '0'
        tenant.correo = 'test@test.com'

    def setUp(self):
        self.tipo_origen = GenDocumentoTipo.objects.create(nombre='Contrato')
        self.tipo_destino = GenDocumentoTipo.objects.create(nombre='Programación')
        self.documento = GenDocumento.objects.create(
            documento_tipo=self.tipo_origen, fecha=date(2026, 1, 1),
        )

    def _detalle(self, fecha_desde, fecha_hasta, documento=None, **overrides):
        datos = {
            'documento': documento or self.documento,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            # Turno diurno de 8h todos los días de la semana por defecto.
            'hora_desde': time(6, 0),
            'hora_hasta': time(14, 0),
            'lunes': True, 'martes': True, 'miercoles': True, 'jueves': True,
            'viernes': True, 'sabado': True, 'domingo': True,
        }
        datos.update(overrides)
        return GenDocumentoDetalle.objects.create(**datos)

    def _generar(self, anio=2026, mes=6):
        return documento_servicio.generar(
            documento_tipo_origen=self.tipo_origen,
            documento_tipo_destino_id=self.tipo_destino.id,
            anio=anio,
            mes=mes,
        )

    def test_acota_fechas_al_mes(self):
        # Arranca a mitad del mes y termina mucho después: se recorta al fin de mes.
        self._detalle(date(2026, 6, 15), date(2026, 12, 30))

        generados = self._generar()

        self.assertEqual(len(generados), 1)
        detalle = generados[0].documentos_detalles_documento_rel.get()
        self.assertEqual(detalle.fecha_desde, date(2026, 6, 15))
        self.assertEqual(detalle.fecha_hasta, date(2026, 6, 30))

    def test_acota_fecha_desde_anterior_al_mes(self):
        # Empezó antes del periodo: fecha_desde se corre al primer día del mes.
        self._detalle(date(2026, 1, 1), date(2026, 12, 30))

        generados = self._generar()

        detalle = generados[0].documentos_detalles_documento_rel.get()
        self.assertEqual(detalle.fecha_desde, date(2026, 6, 1))
        self.assertEqual(detalle.fecha_hasta, date(2026, 6, 30))

    def test_rango_dentro_del_mes_se_conserva(self):
        self._detalle(date(2026, 6, 10), date(2026, 6, 20))

        generados = self._generar()

        detalle = generados[0].documentos_detalles_documento_rel.get()
        self.assertEqual(detalle.fecha_desde, date(2026, 6, 10))
        self.assertEqual(detalle.fecha_hasta, date(2026, 6, 20))

    def test_no_genera_detalle_que_empieza_despues_del_mes(self):
        # Empieza en julio y se genera junio: no aplica.
        self._detalle(date(2026, 7, 1), date(2026, 12, 30))
        # Uno vigente para que el documento sí se genere.
        vigente = self._detalle(date(2026, 6, 1), date(2026, 6, 30))

        generados = self._generar()

        detalles = list(generados[0].documentos_detalles_documento_rel.all())
        self.assertEqual(len(detalles), 1)
        self.assertEqual(detalles[0].documento_detalle_afectado_id, vigente.id)

    def test_no_genera_detalle_que_termino_antes_del_mes(self):
        self._detalle(date(2026, 1, 1), date(2026, 5, 20))
        vigente = self._detalle(date(2026, 6, 1), date(2026, 6, 30))

        generados = self._generar()

        detalles = list(generados[0].documentos_detalles_documento_rel.all())
        self.assertEqual(len(detalles), 1)
        self.assertEqual(detalles[0].documento_detalle_afectado_id, vigente.id)

    def test_documento_sin_detalles_vigentes_no_se_genera(self):
        self._detalle(date(2026, 7, 1), date(2026, 12, 30))

        with self.assertRaises(ValidationError):
            self._generar()

        self.assertEqual(GenDocumento.objects.filter(documento_tipo=self.tipo_destino).count(), 0)
        # El origen tampoco se toca.
        self.documento.refresh_from_db()
        self.assertEqual(self.documento.fecha, date(2026, 1, 1))

    def test_detalle_sin_fechas_aborta(self):
        self._detalle(date(2026, 6, 1), date(2026, 6, 30))
        self._detalle(None, None)

        with self.assertRaises(ValidationError):
            self._generar()

        # Transacción revertida: no quedó nada generado.
        self.assertEqual(GenDocumento.objects.filter(documento_tipo=self.tipo_destino).count(), 0)

    def test_detalle_sin_horario_aborta(self):
        self._detalle(date(2026, 6, 1), date(2026, 6, 30), hora_desde=None, hora_hasta=None)

        with self.assertRaises(ValidationError):
            self._generar()

        self.assertEqual(GenDocumento.objects.filter(documento_tipo=self.tipo_destino).count(), 0)

    def test_calcula_horas_diurnas_todo_el_mes(self):
        # Turno 06:00-14:00 (8h diurnas), todos los días marcados, junio = 30 días.
        self._detalle(date(2026, 6, 1), date(2026, 6, 30))

        generados = self._generar()

        detalle = generados[0].documentos_detalles_documento_rel.get()
        self.assertEqual(detalle.dias, 30)
        self.assertEqual(detalle.horas_diurnas, 240)   # 8h × 30
        self.assertEqual(detalle.horas_nocturnas, 0)
        self.assertEqual(detalle.horas, 240)

    def test_calcula_horas_reparte_diurnas_y_nocturnas(self):
        # Turno 18:00-06:00 (12h): 1h diurna [18-19) + 11h nocturnas, cruza medianoche.
        self._detalle(
            date(2026, 6, 1), date(2026, 6, 30),
            hora_desde=time(18, 0), hora_hasta=time(6, 0),
        )

        generados = self._generar()

        detalle = generados[0].documentos_detalles_documento_rel.get()
        self.assertEqual(detalle.dias, 30)
        self.assertEqual(detalle.horas_diurnas, 30)    # 1h × 30
        self.assertEqual(detalle.horas_nocturnas, 330)  # 11h × 30
        self.assertEqual(detalle.horas, 360)

    def test_cuenta_solo_dias_de_semana_marcados(self):
        # Solo lunes marcados. Junio 2026 tiene 5 lunes (1, 8, 15, 22, 29).
        self._detalle(
            date(2026, 6, 1), date(2026, 6, 30),
            lunes=True, martes=False, miercoles=False, jueves=False,
            viernes=False, sabado=False, domingo=False,
        )

        generados = self._generar()

        detalle = generados[0].documentos_detalles_documento_rel.get()
        self.assertEqual(detalle.dias, 5)
        self.assertEqual(detalle.horas_diurnas, 40)    # 8h × 5

    def test_festivo_manda_sobre_dia_de_semana(self):
        # Lunes marcado pero festivo=False: un lunes festivo no cuenta.
        # Junio 2026: los lunes son 1, 8, 15, 22, 29; marcamos el 15 como festivo.
        GenFestivo.objects.create(id=1, fecha=date(2026, 6, 15), nombre='Festivo')
        self._detalle(
            date(2026, 6, 1), date(2026, 6, 30),
            lunes=True, martes=False, miercoles=False, jueves=False,
            viernes=False, sabado=False, domingo=False, festivo=False,
        )

        generados = self._generar()

        detalle = generados[0].documentos_detalles_documento_rel.get()
        # 5 lunes menos el 15 (festivo, no cobrado) = 4.
        self.assertEqual(detalle.dias, 4)

    def test_festivo_cuenta_cuando_flag_activo(self):
        GenFestivo.objects.create(id=1, fecha=date(2026, 6, 15), nombre='Festivo')
        self._detalle(
            date(2026, 6, 1), date(2026, 6, 30),
            lunes=True, martes=False, miercoles=False, jueves=False,
            viernes=False, sabado=False, domingo=False, festivo=True,
        )

        generados = self._generar()

        detalle = generados[0].documentos_detalles_documento_rel.get()
        # 4 lunes ordinarios + el 15 festivo (festivo=True) = 5.
        self.assertEqual(detalle.dias, 5)


class ModeloPermisoTests(TenantTestCase):
    """
    `GET /general/modelo/<id>/permiso/` decide qué tipos consultan permisos.

    El front pinta los botones de crear/editar/eliminar con esta respuesta, así que la
    lista de tipos que sí se restringen es la que hay que fijar: si alguien agrega un
    tipo nuevo, tiene que decidir a conciencia de qué lado cae.
    """

    @classmethod
    def setup_tenant(cls, tenant):
        tenant.nombre = 'Test permiso'
        tenant.celular = '0'
        tenant.correo = 'permiso@test.com'

    def setUp(self):
        self.factory = APIRequestFactory()
        # Sin `UserTenantPermissions` en este schema, `has_perm` devuelve False para
        # todo: es el usuario recién invitado y sin grupos.
        self.usuario = SegUsuario.objects.create(
            email='sinpermisos@ejemplo.com', is_active=True, is_verified=True,
        )

    def _permiso(self, tipo):
        modelo = GenModelo.objects.create(
            id=90000 + ord(tipo), app='general', clase='GenItem',
            nombre=f'Prueba {tipo}', tabla='gen_item', tipo=tipo,
        )
        peticion = self.factory.get(f'/general/modelo/{modelo.pk}/permiso/')
        force_authenticate(peticion, user=self.usuario)
        vista = _ModeloViewSinPermisos.as_view({'get': 'permiso'})
        return vista(peticion, pk=modelo.pk).data

    def test_los_tipos_restringidos_consultan_permisos(self):
        for tipo in GenModelo.TIPOS_CON_PERMISO:
            with self.subTest(tipo=tipo):
                self.assertEqual(
                    self._permiso(tipo),
                    {'ver': False, 'crear': False, 'editar': False, 'eliminar': False},
                )

    def test_los_demas_tipos_no_se_restringen(self):
        for tipo in (GenModelo.Tipo.FIXTURE, GenModelo.Tipo.DETALLE, GenModelo.Tipo.SOPORTE):
            with self.subTest(tipo=tipo):
                self.assertEqual(
                    self._permiso(tipo),
                    {'ver': True, 'crear': True, 'editar': True, 'eliminar': True},
                )

    def test_movimiento_se_restringe_y_soporte_no(self):
        """
        Explícito porque son las dos decisiones tomadas a mano: Movimiento entró al
        control por permisos, y Soporte se dejó afuera por ser funcionalidad vertical
        (adjuntos, soportes de turno). Cualquiera de las dos se revertiría en silencio.
        """
        self.assertIn(GenModelo.Tipo.MOVIMIENTO, GenModelo.TIPOS_CON_PERMISO)
        self.assertNotIn(GenModelo.Tipo.SOPORTE, GenModelo.TIPOS_CON_PERMISO)


class PermisoDeModeloCoherenteTests(SimpleTestCase):
    """
    La regla, en una sola prueba: un tipo está en `TIPOS_CON_PERMISO` **si y solo si** el
    viewset de sus modelos declara `TienePermisoModelo`.

    Las dos mitades viven en archivos distintos —la taxonomía en `GenModelo`, la
    permission class en cada viewset— y nada obliga a que coincidan. Cuando se separan,
    el front esconde botones que la API acepta, o los muestra y la API responde 403.
    Ambos casos son silenciosos en producción.
    """

    MONTAJES = ['general.urls', 'contabilidad.urls', 'turno.urls',
                'humano.urls', 'inventario.urls', 'seguridad.urls_tenant']

    @staticmethod
    def _modelo_de(viewset):
        qs = getattr(viewset, 'queryset', None)
        if qs is not None:
            return qs.model
        return getattr(getattr(getattr(viewset, 'serializer_class', None), 'Meta', None), 'model', None)

    def _viewsets_con_tipo(self):
        """(etiqueta, tipo declarado en el fixture, ¿declara TienePermisoModelo?)"""
        import importlib
        import json

        with open('general/fixtures/15_modelo.json') as archivo:
            tipos = {r['clase']: r['tipo'] for r in json.load(archivo)['data']}

        for modulo in self.MONTAJES:
            for prefijo, viewset, _ in importlib.import_module(modulo).router.registry:
                modelo = self._modelo_de(viewset)
                if modelo is None or modelo.__name__ not in tipos:
                    continue
                exige = 'TienePermisoModelo' in [c.__name__ for c in viewset.permission_classes]
                yield f'{modulo.split(".")[0]}/{prefijo}', tipos[modelo.__name__], exige

    def test_los_tipos_restringidos_declaran_la_permission_class(self):
        for etiqueta, tipo, exige in self._viewsets_con_tipo():
            if tipo not in GenModelo.TIPOS_CON_PERMISO:
                continue
            with self.subTest(endpoint=etiqueta, tipo=tipo):
                self.assertTrue(exige, f'{etiqueta} es tipo {tipo}: le falta TienePermisoModelo')

    def test_los_tipos_no_restringidos_no_la_declaran(self):
        for etiqueta, tipo, exige in self._viewsets_con_tipo():
            if tipo in GenModelo.TIPOS_CON_PERMISO:
                continue
            with self.subTest(endpoint=etiqueta, tipo=tipo):
                self.assertFalse(exige, f'{etiqueta} es tipo {tipo}: le sobra TienePermisoModelo')


class SubirArchivoTests(TenantTestCase):
    """
    Validación de `POST /general/archivo/`.

    Lo que se prueba no es la subida a B2 (que se mockea) sino que ninguna
    entrada inválida llegue a tocarla: cada request rechazado que hubiera
    llegado a `backblaze.subir` deja un objeto huérfano en el bucket, porque
    la fila que lo referencia nunca se crea.
    """

    @classmethod
    def setup_tenant(cls, tenant):
        tenant.nombre = 'Test'
        tenant.celular = '0'
        tenant.correo = 'test@test.com'

    def setUp(self):
        self.modelo = GenModelo.objects.create(
            id=99001, app='general', clase='GenDocumentoTipo',
            nombre='Documento tipo', tabla='gen_documento_tipo', tipo='F',
        )
        # El schema de prueba nace vacío: el id 1 que usa el default del modelo
        # lo trae `cargar_datos_tenant`, que acá no corre.
        GenArchivoTipo.objects.get_or_create(id=1, defaults={'codigo': 'general', 'nombre': 'General'})
        self.archivo_tipo = GenArchivoTipo.objects.create(
            id=99001, codigo='prueba', nombre='Prueba',
        )
        self.objeto = GenDocumentoTipo.objects.create(nombre='Prueba')
        self.vista = _ArchivoViewSinPermisos.as_view({'post': 'create'})
        self.vista_descargar = _ArchivoViewSinPermisos.as_view({'get': 'descargar'})

    def _pdf(self, nombre='a.pdf'):
        return SimpleUploadedFile(nombre, b'%PDF-1.4 contenido', content_type='application/pdf')

    def _zip(self, interno):
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w') as z:
            z.writestr('[Content_Types].xml', '<x/>')
            z.writestr(interno, 'x')
        return buffer.getvalue()

    def _subir(self, **datos):
        peticion = APIRequestFactory().post('/general/archivo/', datos, format='multipart')
        force_authenticate(peticion, user=SegUsuario(id=1))
        return self.vista(peticion)

    def test_sube_y_crea_la_fila(self):
        with mock.patch.object(backblaze, 'subir') as subir:
            respuesta = self._subir(
                archivo=self._pdf(), modelo=self.modelo.pk, objeto_id=self.objeto.pk,
                archivo_tipo=self.archivo_tipo.pk,
            )
        self.assertEqual(respuesta.status_code, 201)
        self.assertEqual(subir.call_count, 1)
        fila = GenArchivo.objects.get(pk=respuesta.data['id'])
        self.assertEqual(fila.objeto_id, str(self.objeto.pk))
        self.assertEqual(fila.archivo_tipo_id, self.archivo_tipo.pk)
        self.assertEqual(fila.almacenamiento_id, subir.call_args.kwargs['key'])

    def test_archivo_tipo_omitido_toma_el_default(self):
        with mock.patch.object(backblaze, 'subir'):
            respuesta = self._subir(archivo=self._pdf(), modelo=self.modelo.pk, objeto_id=self.objeto.pk)
        self.assertEqual(respuesta.status_code, 201)
        self.assertEqual(GenArchivo.objects.get(pk=respuesta.data['id']).archivo_tipo_id, 1)

    def test_archivo_tipo_vacio_toma_el_default(self):
        """En multipart un campo presente pero vacío llega como '', no ausente."""
        with mock.patch.object(backblaze, 'subir'):
            respuesta = self._subir(
                archivo=self._pdf(), modelo=self.modelo.pk, objeto_id=self.objeto.pk, archivo_tipo='',
            )
        self.assertEqual(respuesta.status_code, 201)
        self.assertEqual(GenArchivo.objects.get(pk=respuesta.data['id']).archivo_tipo_id, 1)

    def test_entradas_invalidas_no_tocan_b2(self):
        casos = {
            'sin archivo': {'modelo': self.modelo.pk, 'objeto_id': self.objeto.pk},
            'sin modelo': {'archivo': self._pdf(), 'objeto_id': self.objeto.pk},
            'sin objeto_id': {'archivo': self._pdf(), 'modelo': self.modelo.pk},
            'modelo inexistente': {'archivo': self._pdf(), 'modelo': 999999, 'objeto_id': self.objeto.pk},
            'modelo no numerico': {'archivo': self._pdf(), 'modelo': 'abc', 'objeto_id': self.objeto.pk},
            'archivo_tipo inexistente': {
                'archivo': self._pdf(), 'modelo': self.modelo.pk,
                'objeto_id': self.objeto.pk, 'archivo_tipo': 999999,
            },
            'objeto_id no entero': {
                'archivo': self._pdf(), 'modelo': self.modelo.pk, 'objeto_id': 'x' * 51,
            },
            'objeto_id fuera del rango de bigint': {
                'archivo': self._pdf(), 'modelo': self.modelo.pk, 'objeto_id': 10 ** 25,
            },
            'objeto_id cero': {
                'archivo': self._pdf(), 'modelo': self.modelo.pk, 'objeto_id': 0,
            },
            'objeto_id negativo': {
                'archivo': self._pdf(), 'modelo': self.modelo.pk, 'objeto_id': -1,
            },
            'objeto_id decimal': {
                'archivo': self._pdf(), 'modelo': self.modelo.pk, 'objeto_id': '1.5',
            },
            'archivo_tipo no entero': {
                'archivo': self._pdf(), 'modelo': self.modelo.pk,
                'objeto_id': self.objeto.pk, 'archivo_tipo': 'abc',
            },
            'tipo no permitido': {
                'archivo': SimpleUploadedFile('a.exe', b'MZ', content_type='application/x-msdownload'),
                'modelo': self.modelo.pk, 'objeto_id': self.objeto.pk,
            },
            'objeto_id inexistente': {
                'archivo': self._pdf(), 'modelo': self.modelo.pk, 'objeto_id': 999999,
            },
            'objeto_id no numerico': {
                'archivo': self._pdf(), 'modelo': self.modelo.pk, 'objeto_id': 'abc',
            },
            'ejecutable disfrazado de pdf': {
                'archivo': SimpleUploadedFile('a.pdf', b'MZ\x90\x00\x03\x00', content_type='application/pdf'),
                'modelo': self.modelo.pk, 'objeto_id': self.objeto.pk,
            },
            'zip disfrazado de docx': {
                'archivo': SimpleUploadedFile('a.docx', self._zip('foto.jpg'), content_type=mime.DOCX),
                'modelo': self.modelo.pk, 'objeto_id': self.objeto.pk,
            },
            'png declarado como pdf': {
                'archivo': SimpleUploadedFile('a.pdf', b'\x89PNG\r\n\x1a\n', content_type='application/pdf'),
                'modelo': self.modelo.pk, 'objeto_id': self.objeto.pk,
            },
        }
        for nombre, datos in casos.items():
            with self.subTest(caso=nombre):
                # Se cuenta antes y después en vez de exigir la tabla vacía: si un
                # caso deja una fila, el siguiente no tiene por qué acusar el error.
                antes = GenArchivo.objects.count()
                with mock.patch.object(backblaze, 'subir') as subir:
                    respuesta = self._subir(**datos)
                self.assertEqual(respuesta.status_code, 400)
                subir.assert_not_called()
                self.assertEqual(GenArchivo.objects.count(), antes)

    def test_archivo_demasiado_grande_no_toca_b2(self):
        # Se baja el techo en vez de fabricar 20 MB: el `size` de un archivo
        # trucado se pierde al re-parsear el multipart.
        with mock.patch.object(archivo_servicio, 'TAMANO_MAXIMO_ARCHIVO', 4), \
             mock.patch.object(backblaze, 'subir') as subir:
            respuesta = self._subir(archivo=self._pdf(), modelo=self.modelo.pk, objeto_id=self.objeto.pk)
        self.assertEqual(respuesta.status_code, 400)
        subir.assert_not_called()

    def test_si_falla_la_fila_se_borra_el_objeto_de_b2(self):
        """Sin esto el objeto queda en el bucket sin nada que lo referencie."""
        with mock.patch.object(backblaze, 'subir'), \
             mock.patch.object(backblaze, 'eliminar') as eliminar, \
             mock.patch.object(GenArchivo.objects, 'create', side_effect=RuntimeError('boom')):
            with self.assertRaises(RuntimeError):
                archivo_servicio.subir_archivo(self._pdf(), modelo=self.modelo, objeto_id=self.objeto.pk)
        self.assertEqual(eliminar.call_count, 1)
        self.assertEqual(eliminar.call_args.args[1].split('/')[-1][-4:], '.pdf')

    def test_el_mensaje_distingue_contenido_desconocido_de_incoherente(self):
        """
        Las dos ramas de `validar_archivo` rechazan lo mismo pero explican
        distinto, y el front muestra el mensaje tal cual.
        """
        desconocido = SimpleUploadedFile('a.pdf', b'MZ\x90\x00\x03\x00', content_type='application/pdf')
        incoherente = SimpleUploadedFile('a.pdf', b'\x89PNG\r\n\x1a\n', content_type='application/pdf')

        with self.assertRaises(ValueError) as c:
            archivo_servicio.validar_archivo(desconocido)
        self.assertIn('ningún tipo permitido', str(c.exception))

        with self.assertRaises(ValueError) as c:
            archivo_servicio.validar_archivo(incoherente)
        self.assertIn('application/pdf', str(c.exception))

    def test_acepta_los_tipos_ambiguos_dentro_de_su_grupo(self):
        """
        doc/xls comparten contenedor OLE2 y txt/csv no tienen firma: en esos dos
        grupos se acepta lo declarado, pero solo dentro del grupo.
        """
        ole2 = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1' + b'\x00' * 40
        casos = [
            (ole2, mime.DOC, True),
            (ole2, mime.XLS, True),
            (ole2, mime.PDF, False),
            (b'nombre,valor\nx,1\n', mime.CSV, True),
            (b'nombre,valor\nx,1\n', mime.TXT, True),
            (b'nombre,valor\nx,1\n', mime.DOCX, False),
        ]
        for contenido, declarado, acepta in casos:
            with self.subTest(declarado=declarado, acepta=acepta):
                archivo = SimpleUploadedFile('a.bin', contenido, content_type=declarado)
                if acepta:
                    self.assertEqual(archivo_servicio.validar_archivo(archivo), declarado)
                else:
                    with self.assertRaises(ValueError):
                        archivo_servicio.validar_archivo(archivo)

    def test_objeto_id_se_rechaza_por_rango_y_no_por_inexistencia(self):
        """
        Cero, negativo y desbordado también fallarían al buscar el registro, así
        que se comprueba el *código* del error: lo tiene que parar el campo.
        """
        for nombre, valor, codigo in (
            ('cero', 0, 'min_value'),
            ('negativo', -1, 'min_value'),
            ('mayor que un bigint', 10 ** 25, 'max_value'),
            ('decimal', '1.5', 'invalid'),
        ):
            with self.subTest(caso=nombre):
                with mock.patch.object(backblaze, 'subir'):
                    respuesta = self._subir(
                        archivo=self._pdf(), modelo=self.modelo.pk, objeto_id=valor,
                    )
                self.assertEqual(respuesta.status_code, 400)
                self.assertEqual(respuesta.data['objeto_id'][0].code, codigo)

    def test_los_identificadores_no_enteros_explican_por_que(self):
        """El mensaje lo muestra el front tal cual, así que no puede hablar de pks."""
        for campo, datos in (
            ('modelo', {'archivo': self._pdf(), 'modelo': 'abc', 'objeto_id': self.objeto.pk}),
            ('objeto_id', {'archivo': self._pdf(), 'modelo': self.modelo.pk, 'objeto_id': 'abc'}),
            ('archivo_tipo', {
                'archivo': self._pdf(), 'modelo': self.modelo.pk,
                'objeto_id': self.objeto.pk, 'archivo_tipo': 'abc',
            }),
        ):
            with self.subTest(campo=campo):
                with mock.patch.object(backblaze, 'subir'):
                    respuesta = self._subir(**datos)
                self.assertEqual(respuesta.status_code, 400)
                self.assertEqual([str(m) for m in respuesta.data[campo]], ['Debe ser un número entero.'])

    def test_si_b2_no_responde_devuelve_502_y_no_deja_fila(self):
        """
        Un corte contra B2 salía como 500 con traceback. No es un error del
        cliente ni un bug: es un tercero que no respondió.
        """
        corte = ConnectionClosedError(endpoint_url='https://s3.us-east-005.backblazeb2.com/x')
        # Se mockea el cliente y no `subir`: la traducción a 502 vive dentro de
        # `subir`, así que mockearla saltaría justo lo que se quiere probar.
        with mock.patch.object(backblaze, '_cliente_s3') as cliente:
            cliente.return_value.put_object.side_effect = corte
            respuesta = self._subir(
                archivo=self._pdf(), modelo=self.modelo.pk, objeto_id=self.objeto.pk,
            )
        self.assertEqual(respuesta.status_code, 502)
        self.assertEqual(respuesta.data['detail'].code, 'error_almacenamiento')
        self.assertFalse(GenArchivo.objects.exists())

    def test_el_corte_de_b2_se_traduce_en_la_capa_de_backblaze(self):
        """
        La traducción vive en `backblaze.subir` y no en la vista, así que
        cualquier llamador la hereda — incluida la subida de foto de perfil.
        """
        corte = ConnectionClosedError(endpoint_url='https://s3.us-east-005.backblazeb2.com/x')
        with mock.patch.object(backblaze, '_cliente_s3') as cliente:
            cliente.return_value.put_object.side_effect = corte
            with self.assertRaises(backblaze.ErrorDeAlmacenamiento) as c:
                backblaze.subir('bucket', 'key', b'x', 'text/plain')
        self.assertEqual(c.exception.status_code, 502)

    def test_si_falla_el_borrado_en_b2_la_fila_no_se_pierde(self):
        """Borrar la fila y dejar el objeto arriba sería perderle el rastro."""
        with mock.patch.object(backblaze, 'subir'):
            respuesta = self._subir(
                archivo=self._pdf(), modelo=self.modelo.pk, objeto_id=self.objeto.pk,
            )
        instancia = GenArchivo.objects.get(pk=respuesta.data['id'])

        corte = ConnectionClosedError(endpoint_url='https://s3.us-east-005.backblazeb2.com/x')
        with mock.patch.object(backblaze, '_cliente_s3') as cliente:
            cliente.return_value.delete_object.side_effect = corte
            with self.assertRaises(backblaze.ErrorDeAlmacenamiento):
                archivo_servicio.eliminar_archivo(instancia)
        self.assertTrue(GenArchivo.objects.filter(pk=instancia.pk).exists())

    def test_la_ruta_en_b2_sigue_el_layout_acordado(self):
        """
        `<cliente>/archivos/<modelo>/<anio>/<mes>/<uuid>.<ext>`.

        La ruta se guarda en la fila, así que cambiarla no rompe lo ya subido —
        pero sí cambia dónde aterriza lo nuevo, y el primer segmento es el que
        separa un contenedor de otro dentro del bucket.
        """
        with mock.patch.object(backblaze, 'subir') as subir:
            respuesta = self._subir(
                archivo=self._pdf(), modelo=self.modelo.pk, objeto_id=self.objeto.pk,
            )
        self.assertEqual(respuesta.status_code, 201)

        key = subir.call_args.kwargs['key']
        cliente_pk, carpeta, modelo_id, anio, mes, nombre = key.split('/')
        ahora = timezone.now()

        self.assertEqual(cliente_pk, str(connection.tenant.pk))
        self.assertEqual(carpeta, 'archivos')
        self.assertEqual(modelo_id, str(self.modelo.pk))
        self.assertEqual(anio, f'{ahora:%Y}')
        self.assertEqual(mes, f'{ahora:%m}')
        self.assertEqual(len(mes), 2)  # con cero adelante, para que ordene
        self.assertTrue(nombre.endswith('.pdf'))
        uuid_lib.UUID(nombre.removesuffix('.pdf'))  # revienta si no es un uuid

        # La fila guarda exactamente la key que se mandó a B2.
        self.assertEqual(GenArchivo.objects.get(pk=respuesta.data['id']).almacenamiento_id, key)

    def test_el_objeto_id_ya_no_aparece_en_la_ruta(self):
        """Se sigue guardando en la fila; en la ruta manda la fecha."""
        with mock.patch.object(backblaze, 'subir') as subir:
            self._subir(archivo=self._pdf(), modelo=self.modelo.pk, objeto_id=self.objeto.pk)
        self.assertNotIn(f'/{self.objeto.pk}/', subir.call_args.kwargs['key'])

    def _fila(self, nombre='a.pdf'):
        """Deja la fila creada sin tocar B2; la subida ya está probada aparte."""
        with mock.patch.object(backblaze, 'subir'):
            respuesta = self._subir(
                archivo=self._pdf(nombre), modelo=self.modelo.pk, objeto_id=self.objeto.pk,
            )
        return GenArchivo.objects.get(pk=respuesta.data['id'])

    def _descargar(self, pk):
        peticion = APIRequestFactory().get(f'/general/archivo/{pk}/descargar/')
        force_authenticate(peticion, user=SegUsuario(id=1))
        return self.vista_descargar(peticion, pk=pk)

    def test_descargar_devuelve_el_contenido_de_b2(self):
        fila = self._fila()
        contenido = b'%PDF-1.4 contenido real'

        with mock.patch.object(backblaze, 'descargar', return_value=contenido) as descargar:
            respuesta = self._descargar(fila.pk)

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.content, contenido)
        self.assertEqual(respuesta['Content-Type'], 'application/pdf')
        # Se pide la key que guarda la fila, no una reconstruida.
        self.assertEqual(descargar.call_args.args[1], fila.almacenamiento_id)

    def test_descargar_usa_el_bucket_privado(self):
        """Si esto apunta al bucket público, la descarga rompe: la key no está ahí."""
        fila = self._fila()
        with mock.patch.object(backblaze, 'descargar', return_value=b'x') as descargar:
            self._descargar(fila.pk)
        self.assertEqual(descargar.call_args.args[0], settings.B2_BUCKET_PRIVADO)

    def test_el_nombre_del_archivo_va_escapado(self):
        """
        El nombre lo eligió quien subió: con comillas o acentos, un header armado
        a mano se rompe o se corta.
        """
        fila = self._fila(nombre='año "raro".pdf')
        with mock.patch.object(backblaze, 'descargar', return_value=b'x'):
            respuesta = self._descargar(fila.pk)

        disposicion = respuesta['Content-Disposition']
        self.assertIn('attachment', disposicion)
        self.assertIn("utf-8''a%C3%B1o%20%22raro%22.pdf", disposicion)
        self.assertNotIn('\n', disposicion)

    def test_si_el_objeto_ya_no_esta_en_b2_devuelve_404(self):
        """404 y no 502: el almacenamiento respondió bien, lo que falta es el archivo."""
        fila = self._fila()
        falta = ClientError({'Error': {'Code': 'NoSuchKey'}}, 'GetObject')
        with mock.patch.object(backblaze, '_cliente_s3') as cliente:
            cliente.return_value.get_object.side_effect = falta
            respuesta = self._descargar(fila.pk)

        self.assertEqual(respuesta.status_code, 404)
        self.assertEqual(respuesta.data['detail'].code, 'archivo_no_encontrado')

    def test_si_b2_no_responde_la_descarga_devuelve_502(self):
        fila = self._fila()
        corte = ConnectionClosedError(endpoint_url='https://s3.us-east-005.backblazeb2.com/x')
        with mock.patch.object(backblaze, '_cliente_s3') as cliente:
            cliente.return_value.get_object.side_effect = corte
            respuesta = self._descargar(fila.pk)

        self.assertEqual(respuesta.status_code, 502)
        self.assertEqual(respuesta.data['detail'].code, 'error_almacenamiento')

    def test_descargar_una_fila_inexistente_devuelve_404_sin_llamar_a_b2(self):
        with mock.patch.object(backblaze, 'descargar') as descargar:
            respuesta = self._descargar(999999)
        self.assertEqual(respuesta.status_code, 404)
        descargar.assert_not_called()


class RededocTests(SimpleTestCase):
    """
    El cliente de api.rededoc.uk. No sale a la red: se reemplaza httpx.request,
    que es el único punto por donde el cliente habla con el mundo.
    """

    def _respuesta(self, status=200, json_datos=None, texto=''):
        respuesta = mock.Mock(status_code=status, text=texto)
        if json_datos is None:
            respuesta.json.side_effect = ValueError('no es json')
        else:
            respuesta.json.return_value = json_datos
        return respuesta

    def test_estado_devuelve_los_datos_del_servicio(self):
        with mock.patch.object(
            rededoc_servicio.httpx, 'request',
            return_value=self._respuesta(200, {'servicio': 'nobelio', 'estado': 'ok'}),
        ) as peticion:
            resultado = rededoc_servicio.Rededoc(url='https://api.rededoc.uk', key='k').estado()

        self.assertEqual(resultado, {'error': False, 'status': 200, 'datos': {'servicio': 'nobelio', 'estado': 'ok'}})
        metodo, url = peticion.call_args.args
        self.assertEqual((metodo, url), ('GET', 'https://api.rededoc.uk/estado/'))

    def test_la_llave_viaja_en_el_header_authorization(self):
        with mock.patch.object(
            rededoc_servicio.httpx, 'request', return_value=self._respuesta(200, {}),
        ) as peticion:
            rededoc_servicio.Rededoc(key='PGwRslxy.secreto').estado()

        headers = peticion.call_args.kwargs['headers']
        self.assertEqual(headers['Authorization'], 'Api-Key PGwRslxy.secreto')

    def test_sin_llave_no_manda_authorization(self):
        """Vale para endpoints públicos como /estado/, pero queda en el log."""
        with mock.patch.object(
            rededoc_servicio.httpx, 'request', return_value=self._respuesta(200, {}),
        ) as peticion:
            rededoc_servicio.Rededoc(key='').estado()

        self.assertNotIn('Authorization', peticion.call_args.kwargs['headers'])

    def test_la_url_base_no_duplica_la_barra(self):
        with mock.patch.object(
            rededoc_servicio.httpx, 'request', return_value=self._respuesta(200, {}),
        ) as peticion:
            rededoc_servicio.Rededoc(url='https://api.rededoc.uk/', key='k').estado()

        self.assertEqual(peticion.call_args.args[1], 'https://api.rededoc.uk/estado/')

    def test_un_error_http_marca_error_y_conserva_el_status(self):
        with mock.patch.object(
            rededoc_servicio.httpx, 'request', return_value=self._respuesta(401, {'detail': 'sin permiso'}),
        ):
            resultado = rededoc_servicio.Rededoc(key='mala').estado()

        self.assertEqual(resultado, {'error': True, 'status': 401, 'datos': {'detail': 'sin permiso'}})

    def test_una_respuesta_que_no_es_json_no_revienta(self):
        """Un 502 del proxy llega en HTML; el cliente lo entrega como mensaje."""
        with mock.patch.object(
            rededoc_servicio.httpx, 'request', return_value=self._respuesta(502, None, '<html>Bad Gateway</html>'),
        ):
            resultado = rededoc_servicio.Rededoc(key='k').estado()

        self.assertTrue(resultado['error'])
        self.assertEqual(resultado['status'], 502)
        self.assertIn('Bad Gateway', resultado['datos']['mensaje'])

    def test_si_no_hay_red_devuelve_status_0_sin_propagar_la_excepcion(self):
        with mock.patch.object(
            rededoc_servicio.httpx, 'request', side_effect=httpx.ConnectError('sin ruta al host'),
        ):
            resultado = rededoc_servicio.Rededoc(key='k').estado()

        self.assertTrue(resultado['error'])
        self.assertEqual(resultado['status'], 0)
        self.assertIn('sin ruta al host', resultado['datos']['mensaje'])

    def test_por_defecto_toma_url_y_llave_de_settings(self):
        with self.settings(REDEDOC_URL='https://api.ejemplo.test', REDEDOC_KEY='llave-de-settings'):
            cliente = rededoc_servicio.Rededoc()
        self.assertEqual(cliente.url, 'https://api.ejemplo.test')
        self.assertEqual(cliente.key, 'llave-de-settings')


class _ParametroViewSinPermisos(GenParametroViewSet):
    """Variante sin auth ni throttle: acá se prueba la vista, no la membresía."""
    authentication_classes = []
    permission_classes = [permissions.AllowAny]
    throttle_classes = []


class _FacturaElectronicaViewSinPermisos(GenFacturaElectronicaViewSet):
    """Sin auth ni permisos: acá se prueba la traducción a HTTP, no la membresía."""
    authentication_classes = []
    permission_classes = [permissions.AllowAny]
    throttle_classes = []


class _ConfiguracionViewSinPermisos(GenConfiguracionViewSet):
    authentication_classes = []
    permission_classes = [permissions.AllowAny]
    throttle_classes = []


class GenParametroViewTests(TenantTestCase):
    """
    La vista de `GenParametro`: se lee entera o por campos, y no se escribe.
    """

    def setUp(self):
        self.factory = APIRequestFactory()
        GenParametro.objects.all().delete()

    def _llamar(self, accion, metodo='get', datos=None, **query):
        vista = _ParametroViewSinPermisos.as_view({metodo: accion})
        peticion = getattr(self.factory, metodo)(f'/general/parametro/{accion}/', datos or query)
        force_authenticate(peticion, user=SegUsuario(id=1))
        return vista(peticion)

    def test_obtener_crea_la_fila_si_el_tenant_todavia_no_la_tiene(self):
        respuesta = self._llamar('obtener')

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.data['id'], 1)
        self.assertIs(respuesta.data['gen_factura_electronica_activa'], False)
        self.assertEqual(GenParametro.objects.count(), 1)

    def test_obtener_dos_veces_no_duplica_la_fila(self):
        self._llamar('obtener')
        self._llamar('obtener')
        self.assertEqual(GenParametro.objects.count(), 1)

    def test_obtener_devuelve_el_valor_guardado(self):
        GenParametro.objects.create(id=1, gen_factura_electronica_activa=True)
        respuesta = self._llamar('obtener')
        self.assertIs(respuesta.data['gen_factura_electronica_activa'], True)

    def test_campos_devuelve_solo_lo_pedido(self):
        GenParametro.objects.create(id=1, gen_factura_electronica_activa=True)
        respuesta = self._llamar('campos', campos='gen_factura_electronica_activa')

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.data, {'gen_factura_electronica_activa': True})

    def test_campos_sin_parametro_devuelve_400(self):
        respuesta = self._llamar('campos')
        self.assertEqual(respuesta.status_code, 400)

    def test_campos_rechaza_una_columna_inexistente(self):
        """Sin esta validación, `campos` sería un .values() con entrada del cliente."""
        respuesta = self._llamar('campos', campos='gen_factura_electronica_activa,inventado')

        self.assertEqual(respuesta.status_code, 400)
        self.assertIn('inventado', respuesta.data['detail'])

    def test_campos_repetidos_no_rompen_la_consulta(self):
        respuesta = self._llamar(
            'campos', campos='gen_factura_electronica_activa,gen_factura_electronica_activa',
        )
        self.assertEqual(respuesta.status_code, 200)

    def test_la_vista_no_expone_ninguna_accion_de_escritura(self):
        """
        El punto entero del modelo: si el front pudiera marcar la activación de
        facturación electrónica, dejaría de ser un hecho verificado.
        """
        acciones = {
            nombre for nombre in dir(GenParametroViewSet)
            if getattr(getattr(GenParametroViewSet, nombre, None), 'mapping', None)
        }
        self.assertEqual(acciones, {'obtener', 'campos'})

        metodos = set()
        for nombre in acciones:
            metodos.update(getattr(GenParametroViewSet, nombre).mapping)
        self.assertEqual(metodos, {'get'})

    def test_el_serializer_no_acepta_escritura(self):
        serializer = GenParametroSerializer(
            GenParametro.objects.create(id=1),
            data={'gen_factura_electronica_activa': True},
            partial=True,
        )
        self.assertTrue(serializer.is_valid(), serializer.errors)
        self.assertEqual(serializer.validated_data, {})


class GenConfiguracionViewTests(TenantTestCase):
    """
    `GenConfiguracion` comparte el mixin con `GenParametro`, pero sí se escribe.
    """

    def setUp(self):
        self.factory = APIRequestFactory()

    def test_obtener_y_campos_siguen_funcionando(self):
        vista = _ConfiguracionViewSinPermisos.as_view({'get': 'obtener'})
        peticion = self.factory.get('/general/configuracion/obtener/')
        force_authenticate(peticion, user=SegUsuario(id=1))
        self.assertEqual(vista(peticion).status_code, 200)

        vista = _ConfiguracionViewSinPermisos.as_view({'get': 'campos'})
        peticion = self.factory.get('/general/configuracion/campos/', {'campos': 'gen_uvt'})
        force_authenticate(peticion, user=SegUsuario(id=1))
        respuesta = vista(peticion)
        self.assertEqual(respuesta.status_code, 200)
        self.assertIn('gen_uvt', respuesta.data)

    def test_actualizar_sigue_escribiendo(self):
        vista = _ConfiguracionViewSinPermisos.as_view({'patch': 'actualizar'})
        peticion = self.factory.patch('/general/configuracion/actualizar/', {'gen_uvt': '47065'})
        force_authenticate(peticion, user=SegUsuario(id=1))
        respuesta = vista(peticion)

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(GenConfiguracion.objects.get(id=1).gen_uvt, Decimal('47065'))


class FacturaElectronicaCrearEmisorTests(TenantTestCase):
    """
    La creación del emisor en rededoc. No sale a la red: se reemplaza el cliente.
    """

    def setUp(self):
        GenParametro.objects.all().delete()
        self.configuracion = GenConfiguracion.objects.get_or_create(id=1)[0]
        # El tenant de pruebas no carga los fixtures, así que el catálogo mínimo
        # se arma acá: ciudad -> estado -> país es la cadena que usa el payload.
        pais = GenPais.objects.create(id=250, nombre='Colombia', codigo='CO')
        estado = GenEstado.objects.create(id=1, nombre='Antioquia', codigo='05', pais=pais)
        ciudad = GenCiudad.objects.create(id=1, nombre='Medellín', codigo='05001', estado=estado)
        identificacion = GenIdentificacion.objects.create(
            id=6, nombre='Número de identificación tributaria', codigo='31',
        )
        tipo_persona = GenTipoPersona.objects.create(id=1, nombre='Jurídica')
        self.configuracion.gen_empresa_razon_social = 'Semantica Digital S.A.S'
        self.configuracion.gen_empresa_numero_identificacion = '901192048'
        self.configuracion.gen_empresa_digito_verificacion = '8'
        self.configuracion.gen_empresa_identificacion = identificacion
        self.configuracion.gen_empresa_tipo_persona = tipo_persona
        self.configuracion.gen_empresa_ciudad = ciudad
        self.configuracion.gen_empresa_direccion = 'Calle 10 # 20-30'
        self.configuracion.gen_empresa_telefono = '3205015059'
        self.configuracion.gen_empresa_correo = 'contacto@ejemplo.com'
        self.configuracion.save()
        self.ciudad = ciudad

    def _cliente(self, crear=None):
        cliente = mock.Mock(spec=rededoc_servicio.Rededoc)
        cliente.crear_emisor.return_value = crear or {
            'error': False, 'status': 201, 'datos': {'id': 77},
        }
        return cliente

    def test_guarda_el_id_del_emisor_y_no_activa_la_facturacion(self):
        """Tener emisor no es estar activo: `gen_factura_electronica_activa` va aparte."""
        cliente = self._cliente()
        parametro = factura_electronica.crear_emisor(cliente=cliente)

        self.assertEqual(parametro.gen_factura_electronica_emisor, 77)
        self.assertEqual(GenParametro.objects.get(id=1).gen_factura_electronica_emisor, 77)
        self.assertIs(GenParametro.objects.get(id=1).gen_factura_electronica_activa, False)

    def test_el_payload_sale_de_la_configuracion_y_no_lleva_cuenta(self):
        """
        `cuenta` la resuelve rededoc a partir de la API key de la integración;
        mandarla sería la única forma de colgar el emisor de otra cuenta.
        """
        cliente = self._cliente()
        factura_electronica.crear_emisor(cliente=cliente)

        payload = cliente.crear_emisor.call_args.args[0]
        self.assertNotIn('cuenta', payload)
        self.assertEqual(payload['razon_social'], 'Semantica Digital S.A.S')
        self.assertEqual(payload['numero_identificacion'], '901192048')
        self.assertEqual(payload['digito_verificacion'], '8')
        self.assertEqual(payload['direccion'], 'Calle 10 # 20-30')
        # Códigos, no PK: las de nuestro catálogo no significan nada en rededoc.
        self.assertEqual(payload['municipio'], '05001')
        self.assertEqual(payload['departamento'], '05')
        self.assertEqual(payload['pais'], 'CO')

    def test_el_nombre_comercial_sale_del_nombre_corto(self):
        self.configuracion.gen_empresa_nombre_corto = 'Semantica'
        self.configuracion.save()
        cliente = self._cliente()
        factura_electronica.crear_emisor(cliente=cliente)

        self.assertEqual(cliente.crear_emisor.call_args.args[0]['nombre_comercial'], 'Semantica')

    def test_si_falta_un_dato_dice_cual_y_no_llama_a_rededoc(self):
        cliente = self._cliente()
        faltantes = {
            'gen_empresa_razon_social': ('', 'razón social'),
            'gen_empresa_numero_identificacion': ('', 'número de identificación'),
            'gen_empresa_identificacion': (None, 'tipo de identificación'),
            'gen_empresa_tipo_persona': (None, 'tipo de organización'),
            'gen_empresa_ciudad': (None, 'ciudad'),
            'gen_empresa_direccion': (None, 'dirección'),
        }
        for campo, (vacio, esperado) in faltantes.items():
            with self.subTest(campo=campo):
                original = getattr(self.configuracion, campo)
                setattr(self.configuracion, campo, vacio)
                self.configuracion.save()

                with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
                    factura_electronica.crear_emisor(cliente=cliente)

                self.assertEqual(caso.exception.status, 400)
                self.assertIn(esperado, caso.exception.cuerpo['detail'])

                setattr(self.configuracion, campo, original)
                self.configuracion.save()

        cliente.crear_emisor.assert_not_called()

    def test_si_el_emisor_ya_existe_es_un_error_que_lo_dice(self):
        """
        La unicidad del NIT la valida rededoc, no nosotros: acá no se consulta
        antes de crear, y su rechazo es el que ve el usuario.
        """
        cliente = self._cliente(crear={
            'error': True, 'status': 400,
            'datos': {'detail': 'Ya existe un emisor con ese número de identificación.'},
        })
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.crear_emisor(cliente=cliente)

        self.assertEqual(caso.exception.cuerpo, {
            'detail': 'Ya existe un emisor con ese número de identificación.',
        })
        self.assertEqual(caso.exception.status, 400)
        self.assertFalse(GenParametro.objects.filter(gen_factura_electronica_emisor__isnull=False).exists())

    def test_un_rechazo_de_rededoc_es_400_y_no_guarda_el_emisor(self):
        cliente = self._cliente(crear={
            'error': True, 'status': 400, 'datos': {'razon_social': ['Requerido']},
        })
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.crear_emisor(cliente=cliente)

        self.assertEqual(caso.exception.status, 400)
        self.assertFalse(GenParametro.objects.filter(gen_factura_electronica_emisor__isnull=False).exists())

    def test_el_error_de_rededoc_sube_tal_cual(self):
        """
        Rededoc responde con la misma forma que nosotros, así que envolverlo dejaba
        al front con el error anidado dentro del error y un mensaje externo que no
        era el de verdad.
        """
        cuerpo = {'detail': 'No se pudo validar el NIT contra el RUES.', 'errores': {}}
        cliente = self._cliente(crear={'error': True, 'status': 503, 'datos': cuerpo})
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.crear_emisor(cliente=cliente)

        self.assertEqual(caso.exception.cuerpo, cuerpo)
        self.assertEqual(caso.exception.status, 502)

    def test_un_rechazo_por_campos_tambien_sube_tal_cual(self):
        cliente = self._cliente(crear={
            'error': True, 'status': 400, 'datos': {'razon_social': ['Requerido']},
        })
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.crear_emisor(cliente=cliente)

        self.assertEqual(caso.exception.cuerpo, {'razon_social': ['Requerido']})
        self.assertEqual(caso.exception.status, 400)

    def test_si_rededoc_no_responde_es_502_y_no_guarda_el_emisor(self):
        cliente = self._cliente(crear={'error': True, 'status': 0, 'datos': {'mensaje': 'timeout'}})
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.crear_emisor(cliente=cliente)

        self.assertEqual(caso.exception.status, 502)
        self.assertFalse(GenParametro.objects.filter(gen_factura_electronica_emisor__isnull=False).exists())


class FacturaElectronicaCertificadoTests(TenantTestCase):
    """
    La carga del certificado. No sale a la red: se reemplaza el cliente.
    """

    def setUp(self):
        GenParametro.objects.all().delete()
        GenParametro.objects.create(id=1, gen_factura_electronica_emisor=77)

    def _archivo(self, nombre='certificado.p12', contenido=b'\x30\x82binario', tamano=None):
        archivo = SimpleUploadedFile(nombre, contenido, content_type='application/x-pkcs12')
        if tamano is not None:
            archivo.size = tamano
        return archivo

    def _cliente(self, respuesta=None):
        cliente = mock.Mock(spec=rededoc_servicio.Rededoc)
        cliente.cargar_certificado.return_value = respuesta or {
            'error': False, 'status': 200, 'datos': {'vigente_hasta': '2027-01-31'},
        }
        return cliente

    def test_manda_emisor_archivo_y_clave_a_rededoc(self):
        cliente = self._cliente()
        datos = factura_electronica.cargar_certificado(self._archivo(), 'secreta', cliente=cliente)

        args, kwargs = cliente.cargar_certificado.call_args
        self.assertEqual(args[0], 77)                      # el emisor guardado, no uno del front
        self.assertEqual(args[2], 'secreta')
        self.assertEqual(kwargs['nombre'], 'certificado.p12')
        self.assertEqual(datos, {'vigente_hasta': '2027-01-31'})

    def test_guarda_el_vencimiento_que_devuelve_rededoc(self):
        factura_electronica.cargar_certificado(self._archivo(), 'secreta', cliente=self._cliente())

        self.assertEqual(
            GenParametro.objects.get(id=1).gen_certificado_vence, date(2027, 1, 31),
        )

    def test_si_rededoc_no_manda_vencimiento_el_campo_queda_vacio(self):
        cliente = self._cliente({'error': False, 'status': 200, 'datos': {}})
        factura_electronica.cargar_certificado(self._archivo(), 'secreta', cliente=cliente)

        self.assertIsNone(GenParametro.objects.get(id=1).gen_certificado_vence)

    def test_un_rechazo_no_toca_el_vencimiento_guardado(self):
        """Si rededoc no aceptó el certificado nuevo, sigue vigente el anterior."""
        GenParametro.objects.filter(id=1).update(gen_certificado_vence=date(2026, 12, 31))
        cliente = self._cliente({'error': True, 'status': 400, 'datos': {'detail': 'Clave mala.'}})
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica):
            factura_electronica.cargar_certificado(self._archivo(), 'mala', cliente=cliente)

        self.assertEqual(
            GenParametro.objects.get(id=1).gen_certificado_vence, date(2026, 12, 31),
        )

    def test_sin_archivo_o_sin_clave_no_llama_a_rededoc(self):
        cliente = self._cliente()
        faltantes = (
            ((None, 'secreta'), 'archivo'),
            ((self._archivo(), ''), 'clave'),
        )
        for (archivo, clave), esperado in faltantes:
            with self.subTest(falta=esperado):
                with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
                    factura_electronica.cargar_certificado(archivo, clave, cliente=cliente)
                self.assertIn(esperado, caso.exception.cuerpo['detail'])
                self.assertEqual(caso.exception.status, 400)

        cliente.cargar_certificado.assert_not_called()

    def test_sin_emisor_no_llama_a_rededoc(self):
        """El certificado se cuelga del emisor: sin emisor no hay dónde ponerlo."""
        GenParametro.objects.filter(id=1).update(gen_factura_electronica_emisor=None)
        cliente = self._cliente()
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.cargar_certificado(self._archivo(), 'secreta', cliente=cliente)

        self.assertIn('crear el emisor', caso.exception.cuerpo['detail'])
        self.assertEqual(caso.exception.status, 400)
        cliente.cargar_certificado.assert_not_called()

    def test_rechaza_una_extension_que_no_es_de_certificado(self):
        cliente = self._cliente()
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.cargar_certificado(
                self._archivo(nombre='contrato.pdf'), 'secreta', cliente=cliente,
            )

        self.assertIn('.p12', caso.exception.cuerpo['detail'])
        cliente.cargar_certificado.assert_not_called()

    def test_acepta_pfx_y_no_distingue_mayusculas(self):
        cliente = self._cliente()
        factura_electronica.cargar_certificado(
            self._archivo(nombre='FIRMA.PFX'), 'secreta', cliente=cliente,
        )
        cliente.cargar_certificado.assert_called_once()

    def test_rechaza_un_archivo_demasiado_grande(self):
        cliente = self._cliente()
        grande = self._archivo(tamano=factura_electronica.TAMANO_MAXIMO_CERTIFICADO + 1)
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.cargar_certificado(grande, 'secreta', cliente=cliente)

        self.assertIn('límite', caso.exception.cuerpo['detail'])
        cliente.cargar_certificado.assert_not_called()

    def test_un_rechazo_de_rededoc_sube_tal_cual(self):
        """Típico: la clave no abre el certificado."""
        cuerpo = {'detail': 'La clave no corresponde al certificado.', 'errores': {}}
        cliente = self._cliente({'error': True, 'status': 400, 'datos': cuerpo})
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.cargar_certificado(self._archivo(), 'mala', cliente=cliente)

        self.assertEqual(caso.exception.cuerpo, cuerpo)
        self.assertEqual(caso.exception.status, 400)

    def test_si_rededoc_no_responde_es_502(self):
        cliente = self._cliente({'error': True, 'status': 0, 'datos': {'mensaje': 'timeout'}})
        with self.assertRaises(factura_electronica.ErrorFacturaElectronica) as caso:
            factura_electronica.cargar_certificado(self._archivo(), 'secreta', cliente=cliente)

        self.assertEqual(caso.exception.status, 502)

    def test_el_cliente_manda_multipart_con_los_tres_campos(self):
        """Contrato con rededoc: `emisor` y `clave` como campos, `archivo` como file."""
        respuesta = {'error': False, 'status': 200, 'datos': {}}
        with mock.patch.object(rededoc_servicio.Rededoc, '_peticion', return_value=respuesta) as peticion:
            rededoc_servicio.Rededoc(key='k').cargar_certificado(77, b'x', 'secreta')

        _, kwargs = peticion.call_args
        self.assertEqual(kwargs['datos'], {'emisor': 77, 'clave': 'secreta'})
        self.assertIn('archivo', kwargs['archivos'])


class FacturaElectronicaVistaTests(TenantTestCase):
    """La vista: solo traduce el servicio a HTTP."""

    def setUp(self):
        self.factory = APIRequestFactory()

    def _llamar(self, vista_clase=None):
        vista = (vista_clase or _FacturaElectronicaViewSinPermisos).as_view({'post': 'crear_emisor'})
        peticion = self.factory.post('/general/factura-electronica/crear-emisor/')
        force_authenticate(peticion, user=SegUsuario(id=1))
        return vista(peticion)

    def test_una_creacion_correcta_responde_200(self):
        parametro = GenParametro(id=1, gen_factura_electronica_emisor=77)
        with mock.patch.object(factura_electronica, 'crear_emisor', return_value=parametro):
            respuesta = self._llamar()

        self.assertEqual(respuesta.status_code, 200)

    def test_traduce_el_error_del_servicio_con_su_status(self):
        error = factura_electronica.ErrorFacturaElectronica(
            {'detail': 'Faltan datos', 'errores': {'campos': ['Razón social']}}, status=400,
        )
        with mock.patch.object(factura_electronica, 'crear_emisor', side_effect=error):
            respuesta = self._llamar()

        self.assertEqual(respuesta.status_code, 400)
        self.assertEqual(respuesta.data, {
            'detail': 'Faltan datos', 'errores': {'campos': ['Razón social']},
        })

    def test_un_mensaje_propio_del_servicio_sale_como_detail(self):
        error = factura_electronica.ErrorFacturaElectronica('Falta la ciudad de la empresa.')
        with mock.patch.object(factura_electronica, 'crear_emisor', side_effect=error):
            respuesta = self._llamar()

        self.assertEqual(respuesta.status_code, 400)
        self.assertEqual(respuesta.data, {'detail': 'Falta la ciudad de la empresa.'})

    def test_un_502_del_servicio_llega_como_502(self):
        error = factura_electronica.ErrorFacturaElectronica('Sin respuesta', status=502)
        with mock.patch.object(factura_electronica, 'crear_emisor', side_effect=error):
            respuesta = self._llamar()
        self.assertEqual(respuesta.status_code, 502)


class _PrecioDetalleViewSinPermisos(GenPrecioDetalleViewSet):
    """Variante sin auth ni throttle: acá se prueba la validación, no la membresía."""
    authentication_classes = []
    permission_classes = [permissions.AllowAny]
    throttle_classes = []


class _PrecioDetalleBaseTests(TenantTestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.precio = GenPrecio.objects.create(
            nombre='Lista 1', fecha_vence=date(2026, 12, 31),
        )
        self.item = GenItem.objects.create(nombre='Item 1')


class CrudPrecioDetalleTests(_PrecioDetalleBaseTests):
    def _crear_detalle(self, **overrides):
        return GenPrecioDetalle.objects.create(**{
            'precio': self.precio,
            'item': self.item,
            'vr_precio': Decimal('100.00'),
            **overrides,
        })

    def test_crea_detalle(self):
        view = _PrecioDetalleViewSinPermisos.as_view({'post': 'create'})
        payload = {'precio': self.precio.id, 'item': self.item.id, 'vr_precio': '150.00'}
        request = self.factory.post('/precio-detalle/', payload, format='json')
        response = view(request)

        self.assertEqual(response.status_code, 201)
        detalle = GenPrecioDetalle.objects.get()
        self.assertEqual(detalle.precio_id, self.precio.id)
        self.assertEqual(detalle.item_id, self.item.id)
        self.assertEqual(detalle.vr_precio, Decimal('150.00'))

    def test_crear_requiere_item(self):
        view = _PrecioDetalleViewSinPermisos.as_view({'post': 'create'})
        payload = {'precio': self.precio.id, 'vr_precio': '150.00'}
        request = self.factory.post('/precio-detalle/', payload, format='json')
        response = view(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn('item', response.data)
        self.assertEqual(GenPrecioDetalle.objects.count(), 0)

    def test_crear_rechaza_precio_item_duplicado(self):
        self._crear_detalle()

        view = _PrecioDetalleViewSinPermisos.as_view({'post': 'create'})
        payload = {'precio': self.precio.id, 'item': self.item.id, 'vr_precio': '999.00'}
        request = self.factory.post('/precio-detalle/', payload, format='json')
        response = view(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn('detail', response.data)
        self.assertEqual(GenPrecioDetalle.objects.count(), 1)

    def test_mismo_item_en_otro_precio_si_se_permite(self):
        self._crear_detalle()
        precio2 = GenPrecio.objects.create(nombre='Lista 2', fecha_vence=date(2026, 12, 31))

        view = _PrecioDetalleViewSinPermisos.as_view({'post': 'create'})
        payload = {'precio': precio2.id, 'item': self.item.id, 'vr_precio': '200.00'}
        request = self.factory.post('/precio-detalle/', payload, format='json')
        response = view(request)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(GenPrecioDetalle.objects.count(), 2)

    def test_actualizar_no_choca_consigo_mismo(self):
        detalle = self._crear_detalle()

        view = _PrecioDetalleViewSinPermisos.as_view({'patch': 'partial_update'})
        request = self.factory.patch('/precio-detalle/', {'vr_precio': '300.00'}, format='json')
        response = view(request, pk=detalle.pk)

        self.assertEqual(response.status_code, 200)
        detalle.refresh_from_db()
        self.assertEqual(detalle.vr_precio, Decimal('300.00'))

    def test_actualizar_rechaza_duplicado(self):
        item2 = GenItem.objects.create(nombre='Item 2')
        self._crear_detalle()
        detalle2 = self._crear_detalle(item=item2)

        # Mover detalle2 al item ya ocupado por el primero dentro del mismo precio.
        view = _PrecioDetalleViewSinPermisos.as_view({'patch': 'partial_update'})
        request = self.factory.patch('/precio-detalle/', {'item': self.item.id}, format='json')
        response = view(request, pk=detalle2.pk)

        self.assertEqual(response.status_code, 400)
        self.assertIn('detail', response.data)
        detalle2.refresh_from_db()
        self.assertEqual(detalle2.item_id, item2.id)

    def test_la_bd_tambien_rechaza_el_duplicado(self):
        self._crear_detalle()

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                GenPrecioDetalle.objects.create(
                    precio=self.precio, item=self.item, vr_precio=Decimal('50.00'),
                )


class ImportarPrecioDetalleTests(_PrecioDetalleBaseTests):
    def _procesar(self, fila):
        serializer = GenPrecioDetalleImportarSerializer()
        # (indice, dict de datos por campo del campos_excel)
        return serializer.procesar_lote([(2, fila)])

    def _fila_valida(self, **overrides):
        return {
            'precio.id': self.precio.id,
            'item.id': self.item.id,
            'vr_precio': '120.50',
            **overrides,
        }

    def test_importa_fila_valida(self):
        creados, errores = self._procesar(self._fila_valida())

        self.assertEqual(errores, [])
        self.assertEqual(creados, 1)
        detalle = GenPrecioDetalle.objects.get()
        self.assertEqual(detalle.precio_id, self.precio.id)
        self.assertEqual(detalle.item_id, self.item.id)
        self.assertEqual(detalle.vr_precio, Decimal('120.50'))

    def test_item_vacio_da_error(self):
        creados, errores = self._procesar(self._fila_valida(**{'item.id': ''}))

        self.assertEqual(creados, 0)
        self.assertEqual(len(errores), 1)
        self.assertIn('Item', errores[0]['mensaje'])
        self.assertEqual(GenPrecioDetalle.objects.count(), 0)

    def test_item_inexistente_da_error(self):
        creados, errores = self._procesar(self._fila_valida(**{'item.id': 9999}))

        self.assertEqual(creados, 0)
        self.assertEqual(len(errores), 1)
        self.assertIn('Item', errores[0]['mensaje'])

    def test_item_es_obligatorio_en_la_plantilla(self):
        self.assertIn('item.id', GenPrecioDetalleImportarSerializer.campos_requeridos)

    def test_duplicado_contra_bd_da_error(self):
        GenPrecioDetalle.objects.create(
            precio=self.precio, item=self.item, vr_precio=Decimal('100.00'),
        )

        creados, errores = self._procesar(self._fila_valida())

        self.assertEqual(creados, 0)
        self.assertEqual(len(errores), 1)
        self.assertIn('Ya existe un detalle', errores[0]['mensaje'])
        self.assertEqual(GenPrecioDetalle.objects.count(), 1)

    def test_duplicado_dentro_del_archivo_da_error(self):
        serializer = GenPrecioDetalleImportarSerializer()
        filas = [(2, self._fila_valida()), (3, self._fila_valida(vr_precio='999.00'))]

        creados, errores = serializer.procesar_lote(filas)

        self.assertEqual(creados, 0)
        self.assertEqual(len(errores), 1)
        self.assertEqual(errores[0]['fila'], 3)
        self.assertIn('Ya existe un detalle', errores[0]['mensaje'])
        self.assertEqual(GenPrecioDetalle.objects.count(), 0)

    def test_mismo_item_en_otro_precio_si_se_importa(self):
        precio2 = GenPrecio.objects.create(nombre='Lista 2', fecha_vence=date(2026, 12, 31))
        serializer = GenPrecioDetalleImportarSerializer()
        filas = [
            (2, self._fila_valida()),
            (3, self._fila_valida(**{'precio.id': precio2.id})),
        ]

        creados, errores = serializer.procesar_lote(filas)

        self.assertEqual(errores, [])
        self.assertEqual(creados, 2)
        self.assertEqual(GenPrecioDetalle.objects.count(), 2)


class ImportarCelularTests(TenantTestCase):
    """
    El importador es la otra puerta de escritura: sin esto entraría por Excel la
    basura que el serializer rechaza en la API. Un celular inválido invalida la fila,
    y el mixin no guarda nada si alguna falla.
    """

    @classmethod
    def setup_tenant(cls, tenant):
        tenant.nombre = 'Test'
        tenant.celular = '0'
        tenant.correo = 'test@test.com'

    def _procesar(self, celular):
        serializer = GenAsesorImportarSerializer()
        return serializer.procesar_lote([
            (2, {'nombre_corto': 'Asesor', 'celular': celular, 'correo': 'a@b.com'}),
        ])

    def test_guarda_el_celular_en_e164(self):
        creados, errores = self._procesar('+57 300 123 4567')

        self.assertEqual(errores, [])
        self.assertEqual(creados, 1)
        self.assertEqual(GenAsesor.objects.get().celular, '+573001234567')

    def test_acepta_un_numero_internacional(self):
        self._procesar('+44 7911 123456')
        self.assertEqual(GenAsesor.objects.get().celular, '+447911123456')

    def test_una_celda_numerica_de_excel_se_rechaza(self):
        """
        Excel devuelve un int cuando el celular se escribió como número, y un número
        pelado no trae indicativo. Para importar, la columna tiene que ser texto y
        traer el `+`; el `00` también sirve y sobrevive mejor a Excel.
        """
        creados, errores = self._procesar(3001234567)

        self.assertEqual(creados, 0)
        self.assertIn('Celular no es válido', errores[0]['mensaje'])

    def test_acepta_el_prefijo_de_salida(self):
        self._procesar('00573001234567')
        self.assertEqual(GenAsesor.objects.get().celular, '+573001234567')

    def test_un_celular_invalido_invalida_la_fila(self):
        creados, errores = self._procesar('123')

        self.assertEqual(creados, 0)
        self.assertEqual(errores[0]['fila'], 2)
        self.assertIn('Celular no es válido', errores[0]['mensaje'])
        self.assertFalse(GenAsesor.objects.exists())

    def test_la_columna_vacia_se_guarda_vacia(self):
        """`celular` no es requerido en la plantilla y el campo no admite null."""
        creados, errores = self._procesar(None)

        self.assertEqual(errores, [])
        self.assertEqual(creados, 1)
        self.assertEqual(GenAsesor.objects.get().celular, '')


class CargarDatosTenantTests(TenantTestCase):
    """
    El cargador corre dentro del request que crea el contenedor, así que su costo es
    latencia de creación: por eso los catálogos van en bloque (`bulk_create`) y no
    fila por fila. Estas pruebas fijan las tres propiedades que ese cambio no puede
    romper: que cargue todo, que las filas con claves distintas no se pisen entre sí,
    y que siga siendo idempotente.
    """

    @classmethod
    def setup_tenant(cls, tenant):
        tenant.nombre = 'Test'
        tenant.celular = '+573000000000'
        tenant.correo = 'test@test.com'

    def setUp(self):
        # `general.signals` cachea en memoria los ids de GenAccion/GenModelo por
        # schema. Acá se siembran esas dos tablas y el rollback del test las vacía,
        # pero la caché no se revierte: sin esto, el primer modelo auditado que
        # guarde cualquier test posterior escribe en `gen_log` una FK que ya no
        # existe. Es la misma limpieza que hace `contenedor/tests.py`.
        from general.signals import limpiar_caches

        self.addCleanup(limpiar_caches)

    def _cargar(self):
        call_command(
            'cargar_datos_tenant',
            schema=self.tenant.schema_name,
            inicial=True,
            stdout=io.StringIO(),
        )

    @staticmethod
    def _fixtures():
        """[(modelo, filas del json)] de todos los archivos que carga el comando."""
        from general.management.commands.cargar_datos_tenant import (
            FIXTURES_DIRS,
            FIXTURES_INICIAL_DIRS,
        )

        archivos = []
        for carpeta in FIXTURES_DIRS + FIXTURES_INICIAL_DIRS:
            archivos += sorted(carpeta.glob('*.json'), key=lambda f: f.name)
        return [json.loads(a.read_text(encoding='utf-8')) for a in archivos]

    def _huella(self):
        """Contenido de todas las tablas sembradas, para comparar entre corridas."""
        huella = {}
        for contenido in self._fixtures():
            modelo = apps.get_model(contenido['model'])
            campos = [f.attname for f in modelo._meta.concrete_fields]
            huella[contenido['model']] = list(
                modelo.objects.order_by('pk').values_list(*campos)
            )
        return huella

    def test_carga_todas_las_filas_de_todos_los_fixtures(self):
        self._cargar()

        for contenido in self._fixtures():
            modelo = apps.get_model(contenido['model'])
            with self.subTest(modelo=contenido['model']):
                self.assertEqual(modelo.objects.count(), len(contenido['data']))

    def test_las_claves_ausentes_no_pisan_a_las_presentes(self):
        """
        `11_documento_tipo.json` tiene diecinueve formas de fila distintas. Al volcar
        en bloque hay que agrupar por conjunto de claves: si se usara la unión, una
        fila que omite una columna la escribiría con el default del modelo y le
        borraría el valor a las que sí la traen.
        """
        self._cargar()

        contenido = next(
            c for c in self._fixtures() if c['model'] == 'general.GenDocumentoTipo'
        )
        for fila in contenido['data']:
            tipo = GenDocumentoTipo.objects.get(pk=fila['id'])
            for clave, esperado in fila.items():
                with self.subTest(id=fila['id'], clave=clave):
                    self.assertEqual(getattr(tipo, clave), esperado)

    def test_es_idempotente_y_no_consulta_por_fila(self):
        self._cargar()
        antes = self._huella()

        with CaptureQueriesContext(connection) as consultas:
            self._cargar()

        self.assertEqual(self._huella(), antes)
        # Medido: 9.000 consultas fila por fila contra 283 en bloque. El tope es
        # holgado a propósito: fija el orden de magnitud, no un número exacto.
        self.assertLess(len(consultas), 500, f'{len(consultas)} consultas')


class _DocumentoDetalleViewSinPermisos(GenDocumentoDetalleViewSet):
    """Variante sin auth ni throttle: acá se prueba la importación, no la membresía."""
    authentication_classes = []
    permission_classes = [permissions.AllowAny]
    throttle_classes = []


class _ImportarDetalleBaseTests(TenantTestCase):
    @classmethod
    def setup_tenant(cls, tenant):
        tenant.nombre = 'Test'
        tenant.celular = '+573000000000'
        tenant.correo = 'test@test.com'

    def setUp(self):
        # Los ids importan: PERFIL_POR_TIPO mapea por documento_tipo_id.
        self.tipo_venta = GenDocumentoTipo.objects.create(id=1, nombre='FACTURA', venta=True)
        self.tipo_compra = GenDocumentoTipo.objects.create(id=5, nombre='COMPRA', compra=True)
        self.tipo_asiento = GenDocumentoTipo.objects.create(id=13, nombre='ASIENTO', contabilidad=True)
        self.tipo_sin_perfil = GenDocumentoTipo.objects.create(id=9, nombre='ENTRADA ALMACEN')

        self.documento = self._documento(self.tipo_venta)
        self.item = GenItem.objects.create(nombre='Servicio')

    def _documento(self, tipo, **overrides):
        return GenDocumento.objects.create(
            documento_tipo=tipo, fecha=date(2026, 1, 15), **overrides,
        )

    def _archivo(self, serializer, filas, nombre='detalles.xlsx'):
        """Arma un .xlsx real con los encabezados que espera la plantilla."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append([
            ImportarExcelMixin._encabezado_importar(campo, encabezado, serializer.campos_requeridos)
            for campo, encabezado in serializer.campos_excel
        ])
        for fila in filas:
            ws.append(fila)
        buf = io.BytesIO()
        wb.save(buf)
        return SimpleUploadedFile(
            nombre, buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _importar(self, filas, documento=None, serializer=None):
        documento = documento or self.documento
        serializer = serializer or GenDocumentoDetalleImportarSerializer(documento)
        archivo = self._archivo(serializer, filas)
        request = APIRequestFactory().post(
            '/general/documento-detalle/importar/',
            {'archivo': archivo, 'documento': documento.id},
            format='multipart',
        )
        return _DocumentoDetalleViewSinPermisos.as_view({'post': 'importar'})(request)

    def _plantilla(self, params):
        request = APIRequestFactory().get('/general/documento-detalle/importar-ejemplo/', params)
        return _DocumentoDetalleViewSinPermisos.as_view({'get': 'importar_ejemplo'})(request)


class ImportarDetallePadreTests(_ImportarDetalleBaseTests):
    """
    El padre es parte del contrato, no un dato del archivo: sin él no hay tipo, y
    sin tipo no hay columnas. Las cuatro puertas se cierran antes de leer el Excel.
    """

    def test_sin_documento_da_400(self):
        response = self._plantilla({})

        self.assertEqual(response.status_code, 400)
        self.assertIn('documento', response.data)

    def test_documento_inexistente_da_404(self):
        response = self._plantilla({'documento': 999999})

        self.assertEqual(response.status_code, 404)

    def test_documento_no_modificable_da_400(self):
        aprobado = self._documento(self.tipo_venta, estado_aprobado=True)

        response = self._plantilla({'documento': aprobado.id})

        self.assertEqual(response.status_code, 400)
        self.assertIn('no es modificable', str(response.data))

    def test_tipo_sin_perfil_da_400(self):
        documento = self._documento(self.tipo_sin_perfil)

        response = self._plantilla({'documento': documento.id})

        self.assertEqual(response.status_code, 400)
        self.assertIn('no admite importación', str(response.data))

    def test_documento_aprobado_entre_validar_y_escribir_no_guarda(self):
        """
        El ViewSet valida el padre antes de leer el archivo; para cuando se
        escribe pudo aprobarse en otra petición. Por eso `procesar_lote` lo
        relee bloqueado en vez de confiar en la instancia que ya tiene.
        """
        serializer = GenDocumentoDetalleImportarSerializer(self.documento)
        GenDocumento.objects.filter(pk=self.documento.pk).update(estado_aprobado=True)

        creados, errores = serializer.procesar_lote([
            (2, {'item.id': self.item.id, 'cantidad': 1, 'precio': 100}),
        ])

        self.assertEqual(creados, 0)
        self.assertIn('no es modificable', errores[0]['mensaje'])
        self.assertFalse(GenDocumentoDetalle.objects.exists())


class ImportarDetallePlantillaTests(_ImportarDetalleBaseTests):
    """La plantilla la decide el tipo del padre: factura y asiento no traen lo mismo."""

    def _encabezados(self, documento):
        from openpyxl import load_workbook

        response = self._plantilla({'documento': documento.id})
        self.assertEqual(response.status_code, 200)
        ws = load_workbook(io.BytesIO(response.content)).active
        return [c.value for c in next(ws.iter_rows())], ws

    def test_plantilla_de_venta_trae_item_y_precio(self):
        encabezados, _ = self._encabezados(self.documento)

        self.assertEqual(encabezados, [
            'Item (ID) *',
            'Cantidad *',
            'Precio *',
            'Porcentaje descuento',
            'Centro de costo (ID)',
            'Detalle',
            'Impuestos separados por coma',
        ])

    def test_el_ejemplo_de_impuestos_muestra_la_lista(self):
        _, ws = self._encabezados(self.documento)

        self.assertEqual(ws.cell(row=2, column=7).value, '1,2')


class ImportarDetalleComercialTests(_ImportarDetalleBaseTests):
    def test_importa_y_recalcula_los_totales_del_padre(self):
        """
        Lo que un `bulk_create` dejaría roto: el detalle sin calcular y el
        encabezado con los totales viejos.
        """
        response = self._importar([
            [self.item.id, 2, 100, None, None, 'Primera', None],
            [self.item.id, 1, 50, 10, None, None, None],
        ])

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data, {'creados': 2})

        detalles = GenDocumentoDetalle.objects.order_by('id')
        self.assertEqual([d.total for d in detalles], [Decimal('200'), Decimal('45')])
        self.assertEqual(detalles[0].detalle, 'Primera')

        self.documento.refresh_from_db()
        self.assertEqual(self.documento.total, Decimal('245'))
        self.assertEqual(self.documento.subtotal, Decimal('250'))
        self.assertEqual(self.documento.descuento, Decimal('5'))

    def test_una_fila_mala_no_deja_nada(self):
        response = self._importar([
            [self.item.id, 2, 100, None, None, None, None],
            [999999, 1, 50, None, None, None, None],
        ])

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data['fase'], 'negocio')
        self.assertEqual(response.data['errores'][0]['fila'], 3)
        self.assertFalse(GenDocumentoDetalle.objects.exists())

        self.documento.refresh_from_db()
        self.assertEqual(self.documento.total, Decimal('0'))

    def test_falta_un_requerido_y_se_reporta_como_estructural(self):
        response = self._importar([[self.item.id, None, 100, None, None, None, None]])

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data['fase'], 'estructural')
        self.assertIn('Cantidad', response.data['errores'][0]['mensaje'])

    def test_aplica_los_impuestos_de_la_columna(self):
        iva = GenImpuesto.objects.create(
            nombre='IVA', nombre_extendido='IVA 19%', porcentaje=19, venta=True, compra=True,
        )

        response = self._importar([[self.item.id, 1, 100, None, None, None, str(iva.id)]])

        self.assertEqual(response.status_code, 200, response.data)
        detalle = GenDocumentoDetalle.objects.get()
        self.assertEqual(detalle.impuesto, Decimal('19'))
        self.assertEqual(detalle.total, Decimal('119'))

    def test_impuesto_que_no_aplica_al_tipo_se_rechaza(self):
        """
        La mitad del contrato que depende del documento_tipo: un impuesto solo de
        compra no puede entrar por la puerta de una factura de venta.
        """
        solo_compra = GenImpuesto.objects.create(
            nombre='RTE', nombre_extendido='Retefuente', porcentaje=2,
            venta=False, compra=True,
        )

        response = self._importar([[self.item.id, 1, 100, None, None, None, str(solo_compra.id)]])

        self.assertEqual(response.status_code, 400)
        self.assertIn('no aplica a documentos de venta', response.data['errores'][0]['mensaje'])
        self.assertFalse(GenDocumentoDetalle.objects.exists())

    def test_el_mismo_impuesto_si_entra_en_una_compra(self):
        solo_compra = GenImpuesto.objects.create(
            nombre='RTE', nombre_extendido='Retefuente', porcentaje=2,
            venta=False, compra=True,
        )
        compra = self._documento(self.tipo_compra)

        response = self._importar(
            [[self.item.id, 1, 100, None, None, None, str(solo_compra.id)]],
            documento=compra,
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(GenDocumentoDetalle.objects.get().impuesto_retencion, Decimal('0'))

    def test_impuesto_inexistente_da_error(self):
        response = self._importar([[self.item.id, 1, 100, None, None, None, '999999']])

        self.assertEqual(response.status_code, 400)
        self.assertIn('Impuesto con id=999999 no existe', response.data['errores'][0]['mensaje'])


class ImportarDetalleContableTests(_ImportarDetalleBaseTests):
    """
    El perfil contable trae las mismas columnas que `ConMovimientoImportarSerializer`
    menos las que pone el padre (comprobante, periodo, fecha, documento). El Excel
    habla de débito y crédito; el modelo guarda naturaleza + valor.
    """

    def setUp(self):
        super().setUp()
        self.asiento = self._documento(self.tipo_asiento)
        self.cuenta = ConCuenta.objects.create(
            codigo='150505', nombre='Terrenos', permite_movimiento=True,
        )
        self.serializer = GenDocumentoDetalleImportarSerializer(self.asiento)

    def _importar_asiento(self, filas):
        return self._importar(filas, documento=self.asiento, serializer=self.serializer)

    def _fila(self, **overrides):
        """Número | Cuenta | Tercero | Centro de costo | Débito | Crédito | Base | Detalle"""
        fila = {
            'numero': 15,
            'cuenta': self.cuenta.id,
            'contacto': None,
            'centro_costo': None,
            'debito': 15000,
            'credito': 0,
            'base': 1000,
            'detalle': 'Ajustes',
        }
        fila.update(overrides)
        return list(fila.values())

    def test_los_encabezados_son_los_del_estandar_contable(self):
        from openpyxl import load_workbook

        response = self._plantilla({'documento': self.asiento.id})

        ws = load_workbook(io.BytesIO(response.content)).active
        self.assertEqual([c.value for c in next(ws.iter_rows())], [
            'Número',
            'Cuenta (ID) *',
            'Tercero (ID)',
            'Centro de costo (ID)',
            'Débito',
            'Crédito',
            'Base',
            'Detalle',
        ])

    def test_el_ejemplo_muestra_un_par_cuadrado(self):
        """Débito y crédito no son campos del modelo: sin ejemplo saldría 'ejemplo 1'."""
        from openpyxl import load_workbook

        response = self._plantilla({'documento': self.asiento.id})

        ws = load_workbook(io.BytesIO(response.content)).active
        debitos = [ws.cell(row=f, column=5).value for f in (2, 3)]
        creditos = [ws.cell(row=f, column=6).value for f in (2, 3)]
        self.assertEqual(debitos, [15000, 0])
        self.assertEqual(creditos, [0, 15000])

    def test_importa_un_asiento(self):
        pais = GenPais.objects.create(id=250, nombre='Colombia', codigo='CO')
        estado = GenEstado.objects.create(id=1, nombre='Antioquia', codigo='05', pais=pais)
        ciudad = GenCiudad.objects.create(id=1, nombre='Medellín', codigo='05001', estado=estado)
        identificacion = GenIdentificacion.objects.create(
            id=6, nombre='Número de identificación tributaria', codigo='31',
        )
        tipo_persona = GenTipoPersona.objects.create(id=1, nombre='Jurídica')
        contacto = GenContacto.objects.create(
            numero_identificacion='123456789', nombre_corto='Tercero', ciudad=ciudad,
            identificacion=identificacion, tipo_persona=tipo_persona,
            direccion='calle 1', telefono='1', correo='t@t.com',
        )
        centro_costo = ConCentroCosto.objects.create(codigo='01', nombre='Administración')

        response = self._importar_asiento([
            self._fila(contacto=contacto.id, centro_costo=centro_costo.id),
            self._fila(numero=16, debito=0, credito=15000, base=0),
        ])

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data, {'creados': 2})

        detalles = GenDocumentoDetalle.objects.order_by('id')
        self.assertEqual([d.numero for d in detalles], [15, 16])
        self.assertEqual([d.naturaleza for d in detalles], ['D', 'C'])
        # El par débito/crédito se traduce a naturaleza + valor, y `cantidad = 1`
        # es lo que hace que `calcular()` deje el valor en total.
        self.assertEqual([d.precio for d in detalles], [Decimal('15000'), Decimal('15000')])
        self.assertEqual([d.cantidad for d in detalles], [Decimal('1'), Decimal('1')])
        self.assertEqual([d.total for d in detalles], [Decimal('15000'), Decimal('15000')])
        self.assertEqual([d.base for d in detalles], [Decimal('1000'), Decimal('0')])
        self.assertEqual(detalles[0].contacto_id, contacto.id)
        self.assertEqual(detalles[0].centro_costo_id, centro_costo.id)
        self.assertEqual(detalles[0].detalle, 'Ajustes')

    def test_credito_guarda_naturaleza_c(self):
        response = self._importar_asiento([self._fila(debito=0, credito=15000)])

        self.assertEqual(response.status_code, 200, response.data)
        detalle = GenDocumentoDetalle.objects.get()
        self.assertEqual(detalle.naturaleza, 'C')
        self.assertEqual(detalle.precio, Decimal('15000'))

    def test_las_celdas_vacias_valen_cero(self):
        """El contador deja en blanco el lado que no usa en vez de escribir 0."""
        response = self._importar_asiento([self._fila(debito=None, credito=15000, base=None)])

        self.assertEqual(response.status_code, 200, response.data)
        detalle = GenDocumentoDetalle.objects.get()
        self.assertEqual(detalle.naturaleza, 'C')
        self.assertEqual(detalle.base, Decimal('0'))

    def test_debito_y_credito_a_la_vez_da_error(self):
        response = self._importar_asiento([self._fila(debito=15000, credito=15000)])

        self.assertEqual(response.status_code, 400)
        self.assertIn('Débito y Crédito a la vez', response.data['errores'][0]['mensaje'])
        self.assertFalse(GenDocumentoDetalle.objects.exists())

    def test_linea_sin_valor_da_error(self):
        response = self._importar_asiento([self._fila(debito=0, credito=0)])

        self.assertEqual(response.status_code, 400)
        self.assertIn('Débito o Crédito mayor que cero', response.data['errores'][0]['mensaje'])

    def test_valor_negativo_da_error(self):
        """Un crédito negativo es un débito mal escrito: el lado lo dice la columna."""
        response = self._importar_asiento([self._fila(debito=0, credito=-15000)])

        self.assertEqual(response.status_code, 400)
        self.assertIn('no puede ser negativo', response.data['errores'][0]['mensaje'])

    def test_cuenta_de_agrupacion_se_rechaza(self):
        """
        Una cuenta que no permite movimiento descuadra los informes del periodo;
        el importador la para acá y no cuando ya está en la BD.
        """
        mayor = ConCuenta.objects.create(codigo='1505', nombre='Terrenos', permite_movimiento=False)

        response = self._importar_asiento([self._fila(cuenta=mayor.id)])

        self.assertEqual(response.status_code, 400)
        self.assertIn('no permite movimientos', response.data['errores'][0]['mensaje'])
        self.assertFalse(GenDocumentoDetalle.objects.exists())

    def test_tercero_inexistente_da_error(self):
        response = self._importar_asiento([self._fila(contacto=999999)])

        self.assertEqual(response.status_code, 400)
        self.assertIn('Tercero con id=999999 no existe', response.data['errores'][0]['mensaje'])

    def test_la_plantilla_contable_no_sirve_para_una_factura(self):
        """Los encabezados se validan contra el perfil del padre, no contra el archivo."""
        archivo = self._archivo(self.serializer, [self._fila()])
        request = APIRequestFactory().post(
            '/general/documento-detalle/importar/',
            {'archivo': archivo, 'documento': self.documento.id},  # documento de venta
            format='multipart',
        )

        response = _DocumentoDetalleViewSinPermisos.as_view({'post': 'importar'})(request)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data['fase'], 'encabezados')
